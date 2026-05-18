from flask import (
    Blueprint,
    request,
    jsonify,
    current_app,
    render_template,
    redirect,
    url_for,
    flash,
    session,
    send_file,
)
import traceback

from werkzeug.security import generate_password_hash, check_password_hash
from app import db
from app.services.pdf_service import PDFService
from app.utils.exceptions import PDFServiceError, URLValidationError
from marshmallow import Schema, fields, ValidationError
import os
from sqlalchemy import desc
from sqlalchemy import delete, update
from app.utils.extract_clause_helper import check_free_report_used

# model import
from app.routes.audit.view import check_evidence_staleness
from app.models.organization import *
from app.models.user import *
from app.models.download import *
from app.models.ai import *
from app.models.auditOrganization import *
from datetime import datetime
import json
from flask_login import login_user, logout_user, login_required, current_user
from app import login_manager
from app.utils.permission_handler import role_required
from app.routes.retrival import *
from app.routes.audit.view import calculate_clause_compliance_status
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from app.utils.permission_handler import *
from app import create_app
from sqlalchemy import select, inspect
from werkzeug.utils import secure_filename
from sqlalchemy.orm import relationship, Mapped, mapped_column, joinedload
from app.utils.cleaning import *
from app.models.project_instance_models import *
from app.services.prompt_service import *
from app.utils.bread_crumb import add_to_breadcrumb
from app.helper.evidence_helper import *
from app.utils.compliance_utils import (
    evaluate_single_activity_ai,
    get_clause_compliance_status,
    get_compliance_status_display_info,
    get_assessment_status,
    get_project_clause_statistics,
    get_project_severity_statistics,
  
    get_project_evidence_statistics,
)

from app.utils.cleaning import get_clause_evidence_availability
from app.services.automate_task import (
    _ci_get,
    _as_json,
    _as_dict,
    _safe_vec_id,
    session_scope,
)
from app.services.manual_task import generate_single_clause_activities
from app.services.prompt_templates.test_procedure import *
from app.services.model_response import *
from app.services.prompt_templates.extract_evidence import *
from app.services.manual_task import process_test_procedures
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from flask import send_from_directory, url_for
from pathlib import Path
from app.utils.cleaning import get_activities_for_clause_table, add_activities_table_to_word

re_bp = Blueprint(
    "re",
    __name__,
    template_folder="../../templates/dashboards/re",
    static_folder="../../templates/dashboards/re/assets",
)


@re_bp.route("/", methods=["GET"])
@role_required("RE")
def dashboard():
    """
    Dashboard route for the RE application.
    """
    try:
        add_to_breadcrumb(request.full_path, "Dashboard")
        # pdf_service = PDFService()
        # guidelines = pdf_service.get_guidelines()
        print(current_user)
        print(current_user.organization_id, current_user.name, current_user.role.name)
        return render_template("dash.html")
    except PDFServiceError as pdf_err:
        current_app.logger.error(f"PDF Service Error: {str(pdf_err)}")
        return jsonify({"error": "Error with PDF service"}), 500
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/add", methods=["GET"])
# @role_required()
def add():
    """
    Add
    """
    try:
        # pdf_service = PDFService()
        # guidelines = pdf_service.get_guidelines()
        return render_template("add.html")
    except PDFServiceError as pdf_err:
        current_app.logger.error(f"PDF Service Error: {str(pdf_err)}")
        return jsonify({"error": "Error with PDF service"}), 500
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/guidelines", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def guidelines():
    """
    Endpoint to get the guidelines for PDF scanning.
    """
    add_to_breadcrumb(request.full_path, "Guidelines")
    try:
        guidelines_query = Guidelines.query
        if current_user.is_authenticated and current_user.auditor_profile_id:
            add_to_breadcrumb(request.full_path, "My Guidelines")
            # Subquery to get all guideline_ids already selected
            subquery = select(auditor_selected_guidelines.c.guideline_id).where(
                auditor_selected_guidelines.c.audit_id
                == current_user.auditor_profile_id
            )

            # Main query: get Guidelines NOT in the subquery
            stmt = select(Guidelines).where(
                ~Guidelines.id.in_(subquery), Guidelines.enabled == True
            )

            guidelines = db.session.execute(stmt).scalars().all()
            current_app.logger.info("If block executed - Auditor view")

            # Check if download was successful from query parameter
            download_success = request.args.get("download_success")
        else:
            # For other roles (COMPLIFYRE, RE) show all guidelines (including disabled)
            guidelines = guidelines_query.order_by(Guidelines.created_at.desc()).all()
            current_app.logger.info("Else block executed - Non-auditor view")
            download_success = None

        # Log guideline count
        current_app.logger.info(f"Retrieved {len(guidelines)} guidelines")

        # Convert guidelines to list of dictionaries for detailed logging
        for i, guideline in enumerate(guidelines):
            current_app.logger.info(f"Guideline {i+1}: {guideline.__dict__}")
        return render_template("view.html", guidelines=guidelines)
    except PDFServiceError as pdf_err:
        current_app.logger.error(f"PDF Service Error: {str(pdf_err)}")
        return jsonify({"error": "Error with PDF service"}), 500
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/guideline/<int:guideline_id>/toggle_enabled", methods=["POST"])
@role_required("COMPLIFYRE")
@login_required
def toggle_guideline_enabled(guideline_id):
    try:
        data = request.get_json(silent=True) or {}
        # optional: client may pass {"enabled": true/false}; otherwise toggle
        enabled_from_client = data.get("enabled", None)
        guideline = Guidelines.query.get(guideline_id)
        if not guideline:
            return jsonify({"error": "Guideline not found"}), 404

        # Compute new value
        if enabled_from_client is None:
            new_enabled = not guideline.enabled
        else:
            new_enabled = bool(enabled_from_client)

        guideline.enabled = new_enabled
        db.session.add(guideline)

        # If disabling, remove associations to auditor_selected_guidelines so auditors won't see it
        if not new_enabled:
            # Assuming auditor_selected_guidelines is a Table object available in scope
            stmt = delete(auditor_selected_guidelines).where(
                auditor_selected_guidelines.c.guideline_id == guideline_id
            )
            db.session.execute(stmt)

        db.session.commit()
        # 🔑 Instead of flash(), return message in JSON
        msg = f"Guideline {guideline.id} has been {'enabled Guideline ✅' if new_enabled else 'disabled Guideline ❌'}"
        return (
            jsonify(
                {"guideline_id": guideline_id, "enabled": new_enabled, "message": msg}
            ),
            200,
        )

    except Exception as e:
        current_app.logger.exception("Error toggling guideline enabled")
        db.session.rollback()
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/login", methods=["GET"])
def login():
    """
    login page
    """
    try:
        # pdf_service = PDFService()
        # guidelines = pdf_service.get_guidelines()
        return render_template("login.html")
    except PDFServiceError as pdf_err:
        current_app.logger.error(f"PDF Service Error: {str(pdf_err)}")
        return jsonify({"error": "Error with PDF service"}), 500
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@login_manager.user_loader
def load_user(user_id):
    """
    Load user based on user_id string.
    Handles both regular Users and OrganizationContacts.
    """
    print(f"=== LOAD_USER CALLED ===")
    print(f"User ID received: {user_id} (type: {type(user_id)})")

    try:
        # Check if it's an OrganizationContact (starts with 'contact_')
        if user_id.startswith("contact_"):
            contact_id = user_id.replace("contact_", "")
            print(f"Loading OrganizationContact with ID: {contact_id}")

            from app.models.organization import OrganizationContacts

            contact = OrganizationContacts.query.get(int(contact_id))

            if contact:
                print(
                    f"OrganizationContact loaded: {contact.name} (ID: {contact.contact_id})"
                )
            else:
                print(f"OrganizationContact not found with ID: {contact_id}")

            return contact
        else:
            # Regular user - pure integer ID
            print(f"Loading regular User with ID: {user_id}")
            user = Users.query.get(int(user_id))

            if user:
                print(f"User loaded: {user.email} (ID: {user.id})")
            else:
                print(f"User not found with ID: {user_id}")

            return user

    except Exception as e:
        print(f"Error in load_user: {str(e)}")
        return None


@re_bp.route("/login-user", methods=["POST"])
def login_user_route():
    """
    Login logic for users.
    """
    try:
        # Extract login credentials from the request
        email = request.form.get("email")
        password = request.form.get("password")
        # Fetch the user from the database
        user = Users.query.filter_by(email=email).first()
        if not user:
            flash("Invalid email or password.", "danger")
            return redirect(url_for("re.login"))

        # Verify the password
        if not check_password_hash(user.password_hash, password):
            flash("Invalid email or password.", "danger")
            return redirect(url_for("re.login"))

        # Check if the user is active
        if user.status != "active":
            print("here")
            flash("Your account is not active. Please contact support.", "warning")
            return redirect(url_for("re.login"))

        # Update last login timestamp
        user.last_login = datetime.utcnow()
        db.session.commit()

        # Log the user in
        login_user(user)
        print(current_user.organization_id, current_user.name, current_user.role)
        if (
            current_user.role.name == "AUDITOR"
            or current_user.role.name == "ADMIN"
            and current_user.auditor_profile_id
        ):
            flash("Login successful!", "success")
            return redirect(url_for("audit.dashboard"))
        elif (
            current_user.role.name == "ADMIN" or current_user.role.name == "COMPLIFYRE"
        ):
            flash("Login successful!", "success")
            return redirect(url_for("main.home"))
        flash("Login successful!", "success")
        return redirect(url_for("re.dashboard"))

    except Exception as e:
        current_app.logger.error(f"Error during login: {str(e)}")
        flash("An error occurred during login. Please try again.", "danger")
        return redirect(url_for("re.login"))


@re_bp.route("/logout", methods=["GET"])
@login_required
def logout():
    """
    Logout logic for users.
    """
    try:
        logout_user()
        flash("You have been logged out.", "success")
        return redirect(url_for("re.login"))
    except Exception as e:
        current_app.logger.error(f"Error during logout: {str(e)}")
        flash("An error occurred during logout. Please try again.", "danger")
        return redirect(url_for("re.dashboard"))


@re_bp.route("/control-cyber", methods=["GET"])
# @role_required()
def cyber():
    """
    View page
    """
    try:
        # pdf_service = PDFService()
        # guidelines = pdf_service.get_guidelines()
        return render_template("view1.html")

    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/admin")
# # @role_required()
def admin():
    """
    Admin Page
    """
    try:
        # Fetch all departments from the database
        departments = (
            db.session.query(OrganizationDepartments.department_name).distinct().all()
        )

        roles = Roles.query.all()
        return render_template("admin_1.html", departments=departments, roles=roles)
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/applicable")
# @role_required()
def applicable():
    """
    Applicable Page
    """
    try:
        return render_template("applicable.html")
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/audit")
# @role_required()
def audit():
    """
    Audit Page
    """
    try:
        return render_template("audit.html")
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/compliance")
# @role_required()
def compliance():
    """
    Compliance Page
    """
    try:
        return render_template("compliance.html")
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/orgnization")
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def orgnization():
    """
    Organization Page
    """
    add_to_breadcrumb(request.full_path, "Orginization")
    try:
        # For COMPLIFYRE users, show ALL organizations
        if hasattr(current_user, "role") and current_user.role == "COMPLIFYRE":
            organizations = Organizations.query.all()
        # For AUDITOR users, show only their clients
        elif current_user and current_user.auditor_profile_id:
            add_to_breadcrumb(request.full_path, "My Clients")
            stmt = select(auditor_client).where(
                auditor_client.c.audit_id == current_user.auditor_profile_id
            )
            results = db.session.execute(stmt).fetchall()
            organization_ids = [client.client_id for client in results]
            organizations = Organizations.query.filter(
                Organizations.organization_id.in_(organization_ids)
            ).all()
        else:
            organizations = []

        print(f"DEBUG: Found {len(organizations)} organizations")
        return render_template("dashboards/re/org.html", organizations=organizations)
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/check-client-delete/<int:org_id>", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def check_client_delete(org_id):
    """Check if client can be deleted"""
    try:
        # Check if client exists
        client = Organizations.query.get(org_id)
        if not client:
            return jsonify({"error": "Client not found"}), 404

        # Check if client has any projects
        project_count = Projects.query.filter_by(client=org_id).count()

        return jsonify(
            {
                "can_delete": project_count == 0,
                "project_count": project_count,
                "client_name": client.legal_name or client.name,
                "message": (
                    f"Cannot delete client as it has {project_count} related project(s)."
                    if project_count > 0
                    else "Client can be deleted"
                ),
            }
        )
    except Exception as err:
        current_app.logger.error(f"Error checking client delete: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/delete-client/<int:org_id>", methods=["DELETE"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def delete_client(org_id):
    """Delete client if no projects exist"""
    try:
        # Check if client exists
        client = Organizations.query.get(org_id)
        if not client:
            return jsonify({"success": False, "message": "Client not found"}), 404

        # Double-check if client has any projects before deletion
        project_count = Projects.query.filter_by(client=org_id).count()
        if project_count > 0:
            return (
                jsonify(
                    {
                        "success": False,
                        "message": f"This client cannot be deleted as it has {project_count} related project(s).",
                    }
                ),
                400,
            )

        # Check if current user is authorized to delete this client
        if current_user.auditor_profile_id:
            stmt = select(auditor_client).where(
                auditor_client.c.audit_id == current_user.auditor_profile_id,
                auditor_client.c.client_id == org_id,
            )
            results = db.session.execute(stmt).fetchall()
            if not results:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "You are not authorized to delete this client",
                        }
                    ),
                    403,
                )

        # Delete related data - IMPORTANT: Order matters due to foreign key constraints
        from app.models.organization import OrganizationAddresses, OrganizationInfo

        # FIRST: Delete from department_in_org table (this is a db.Table() object)
        from app.models.organization import department_in_org

        db.session.execute(
            department_in_org.delete().where(
                department_in_org.c.organization_id == org_id
            )
        )

        # Delete addresses
        OrganizationAddresses.query.filter_by(organization_id=org_id).delete()

        # Delete organization info
        OrganizationInfo.query.filter_by(organization_id=org_id).delete()

        # Delete from auditor_client table
        db.session.execute(
            auditor_client.delete().where(auditor_client.c.client_id == org_id)
        )

        # Finally delete the organization
        db.session.delete(client)
        db.session.commit()

        current_app.logger.info(
            f"Client {org_id} deleted successfully by user {current_user.id}"
        )

        return jsonify(
            {
                "success": True,
                "message": f"Client '{client.legal_name or client.name}' deleted successfully",
            }
        )

    except Exception as err:
        db.session.rollback()
        current_app.logger.error(f"Error deleting client {org_id}: {str(err)}")
        return (
            jsonify(
                {"success": False, "message": f"Error deleting client: {str(err)}"}
            ),
            500,
        )


@re_bp.route("/process")
# @role_required()
def process():
    """
    Process Page
    """
    try:
        # compliance_activities = ComplianceActivities.query.all()
        return render_template("process.html")
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/add-new-client", methods=["GET", "POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def add_new_client():
    """
    Handle adding new client with multi-step form
    """

    if request.method == "POST":
        try:
            current_step = request.form.get("current_step", "0")

            # Handle different steps
            if current_step == "0":
                # Step 1: Organization Basic Info
                org_name = request.form.get("org_name", "").strip()
                legal_name = request.form.get("legal_name", "").strip()
                constitution = request.form.get("constitution", "").strip()
                selected_org_types = request.form.getlist("selected_org_types")
                selected_industries = request.form.getlist("selected_industries")
                indian_regulatory = request.form.get("indian_regulatory", "").strip()

                # Store in session for back navigation
                session["step_0_data"] = {
                    "org_name": org_name,
                    "legal_name": legal_name,
                    "constitution": constitution,
                    "selected_org_types": selected_org_types,
                    "selected_industries": selected_industries,
                    "indian_regulatory": indian_regulatory,
                    "address": request.form.get("address", "").strip(),
                    "country": request.form.get("country", "").strip(),
                    "state": request.form.get("state", "").strip(),
                    "city": request.form.get("city", "").strip(),
                }

                # Check if organization with same name already exists
                # existing_org = Organizations.query.filter_by(name=org_name).first()
                # if existing_org:
                #     flash(f"Organization with ID '{org_name}' already exists.", "error")
                #     return redirect(url_for("re.add_new_client", step=0))

                # Create new organization
                organization = Organizations(
                    name=org_name,
                    legal_name=legal_name,
                    constutution=constitution,
                    organization_type=selected_org_types,
                    industry_type=selected_industries,
                    status="active",
                )
                db.session.add(organization)
                db.session.flush()

                # Handle head office address
                addr_line1 = request.form.get("address", "").strip()
                country = request.form.get("country", "").strip()
                state = request.form.get("state", "").strip()
                city = request.form.get("city", "").strip()

                if addr_line1 and city and state and country:
                    head_office = OrganizationAddresses(
                        organization_id=organization.organization_id,
                        address_type="headoffice",
                        address_line1=addr_line1,
                        city=city,
                        state=state,
                        country=country,
                        is_primary=True,
                    )
                    db.session.add(head_office)

                db.session.commit()

                # Create auditor-client relationship
                if current_user and current_user.auditor_profile_id:
                    stmt = auditor_client.insert().values(
                        audit_id=current_user.auditor_profile_id,
                        client_id=organization.organization_id,
                    )
                    db.session.execute(stmt)

                db.session.commit()

                # Store org_id in session
                session["new_client_org_id"] = organization.organization_id
                session["new_client_org_name"] = organization.name

                flash(
                    "Organization information saved successfully. Please continue to Step 2.",
                    "success",
                )
                return redirect(url_for("re.add_new_client", step=1))

            elif current_step == "1":
                # Step 2: Key Locations
                org_id = session.get("new_client_org_id")
                if not org_id:
                    flash("Please start from Step 1", "error")
                    return redirect(url_for("re.add_new_client"))

                locations_data_str = request.form.get("locations_data", "[]")
                try:
                    locations = json.loads(locations_data_str)
                except json.JSONDecodeError:
                    locations = []

                # Store in session
                session["step_1_data"] = {"locations": locations}

                for loc in locations:
                    if all(k in loc for k in ["address", "country", "state", "city"]):
                        new_address = OrganizationAddresses(
                            organization_id=org_id,
                            address_type="keylocation",
                            address_line1=loc.get("address", ""),
                            city=loc.get("city", ""),
                            state=loc.get("state", ""),
                            country=loc.get("country", ""),
                        )
                        db.session.add(new_address)

                db.session.commit()
                flash("Key locations saved successfully.", "success")
                return redirect(url_for("re.add_new_client", step=2))

            elif current_step == "2":
                # Step 3: Departments
                org_id = session.get("new_client_org_id")
                if not org_id:
                    flash("Please start from Step 1", "error")
                    return redirect(url_for("re.add_new_client"))

                selected_ids = request.form.getlist("department_ids")
                selected_department_ids = [
                    int(x) for x in selected_ids if x and x.isdigit()
                ]

                # Store in session
                session["step_2_data"] = {"department_ids": selected_department_ids}

                if selected_department_ids:
                    new_relationships = [
                        {"organization_id": org_id, "department_id": dept_id}
                        for dept_id in selected_department_ids
                    ]
                    insert_stmt = department_in_org.insert().values(new_relationships)
                    db.session.execute(insert_stmt)

                db.session.commit()
                flash("Departments saved successfully.", "success")
                return redirect(url_for("re.add_new_client", step=3))

            elif current_step == "3":
                # Step 4: Business Overview
                org_id = session.get("new_client_org_id")
                if not org_id:
                    flash("Please start from Step 1", "error")
                    return redirect(url_for("re.add_new_client"))

                business_description = request.form.get(
                    "businessDescription", ""
                ).strip()
                branches_india = request.form.get("branchesIndia", type=int) or 0
                branches_outside_india = (
                    request.form.get("branchesOutsideIndia", type=int) or 0
                )
                business_processes = request.form.get("businessProcesses", "").strip()
                org_history = request.form.get("orgHistory", "").strip()

                # Store in session
                session["step_3_data"] = {
                    "business_description": business_description,
                    "branches_india": branches_india,
                    "branches_outside_india": branches_outside_india,
                    "business_processes": business_processes,
                    "org_history": org_history,
                }

                if not business_description:
                    flash("Business description is required.", "error")
                    return redirect(url_for("re.add_new_client", step=3))

                org_info = OrganizationInfo(
                    organization_id=org_id,
                    business_desc=business_description,
                    history=org_history or "",
                    no_of_branches_in_india=branches_india,
                    no_of_branches_outside_india=branches_outside_india,
                    is_active=True,
                )
                db.session.add(org_info)

                if business_processes:
                    compliance_profile = OrganizationComplianceProfiles(
                        organization_id=org_id, business_process=business_processes
                    )
                    db.session.add(compliance_profile)

                db.session.commit()
                flash("Business overview saved successfully.", "success")
                return redirect(url_for("re.add_new_client", step=4))

            elif current_step == "4":
                # Step 5: Organization Structure
                org_id = session.get("new_client_org_id")
                if not org_id:
                    flash("Please start from Step 1", "error")
                    return redirect(url_for("re.add_new_client"))

                positions_data_str = request.form.get("positions_data", "[]")
                try:
                    positions = json.loads(positions_data_str)
                except json.JSONDecodeError:
                    positions = []

                # Store in session
                session["step_4_data"] = {"positions": positions}

                for position_data in positions:
                    if all(k in position_data for k in ["position", "reportsTo"]):
                        new_structure = OrganizationStructure(
                            organization_id=org_id,
                            position=position_data["position"],
                            report_to=position_data["reportsTo"],
                        )
                        db.session.add(new_structure)

                db.session.commit()
                flash("Organization structure saved successfully.", "success")
                return redirect(url_for("re.add_new_client", step=5))

            elif current_step == "5":
                # Step 6: Financial Overview
                org_id = session.get("new_client_org_id")
                if not org_id:
                    flash("Please start from Step 1", "error")
                    return redirect(url_for("re.add_new_client"))

                financial_data = {
                    "key_revenue": request.form.get("revenueStreams", "").strip(),
                    "key_markets_customers": request.form.get(
                        "marketsCustomers", ""
                    ).strip(),
                    "key_financials": request.form.get("financialMetrics", "").strip(),
                    "total_revenue": request.form.get("totalRevenue", "").strip(),
                    "net_profit_loss": request.form.get("netProfitLoss", "").strip(),
                    "total_assets": request.form.get("totalAssets", "").strip(),
                    "total_liabilities": request.form.get(
                        "totalLiabilities", ""
                    ).strip(),
                    "key_financial_challenges": request.form.get(
                        "financialChallenges", ""
                    ).strip(),
                    "auditors_insights": request.form.get(
                        "auditorsInsights", ""
                    ).strip(),
                    "compliance_status": request.form.get(
                        "complianceStatus", ""
                    ).strip(),
                    "pending_litigations": request.form.get(
                        "pendingLitigations", ""
                    ).strip(),
                    "regulatory_filings": request.form.get(
                        "regulatoryFilings", ""
                    ).strip(),
                }

                # Store in session
                session["step_5_data"] = financial_data

                # Update or create OrganizationInfo
                org_info = OrganizationInfo.query.filter_by(
                    organization_id=org_id, is_active=True
                ).first()

                if org_info:
                    org_info.key_revenue = (
                        financial_data["key_revenue"] or org_info.key_revenue
                    )
                    org_info.key_markets_customers = (
                        financial_data["key_markets_customers"]
                        or org_info.key_markets_customers
                    )
                    org_info.key_financials = (
                        financial_data["key_financials"] or org_info.key_financials
                    )
                    org_info.total_revenue_last_year = (
                        financial_data["total_revenue"]
                        or org_info.total_revenue_last_year
                    )
                    org_info.net_profit_loss = (
                        financial_data["net_profit_loss"] or org_info.net_profit_loss
                    )
                    org_info.total_assets = (
                        financial_data["total_assets"] or org_info.total_assets
                    )
                    org_info.total_liabilities = (
                        financial_data["total_liabilities"]
                        or org_info.total_liabilities
                    )
                    org_info.key_financial_challenges = (
                        financial_data["key_financial_challenges"]
                        or org_info.key_financial_challenges
                    )
                else:
                    org_info = OrganizationInfo(
                        organization_id=org_id,
                        business_desc="",
                        history="",
                        key_revenue=financial_data["key_revenue"],
                        key_markets_customers=financial_data["key_markets_customers"],
                        key_financials=financial_data["key_financials"],
                        total_revenue_last_year=financial_data["total_revenue"],
                        net_profit_loss=financial_data["net_profit_loss"],
                        total_assets=financial_data["total_assets"],
                        total_liabilities=financial_data["total_liabilities"],
                        key_financial_challenges=financial_data[
                            "key_financial_challenges"
                        ],
                        is_active=True,
                    )
                    db.session.add(org_info)

                # Update or create compliance profile
                compliance_profile = OrganizationComplianceProfiles.query.filter_by(
                    organization_id=org_id
                ).first()

                if compliance_profile:
                    compliance_profile.compliance_status = (
                        financial_data["compliance_status"]
                        or compliance_profile.compliance_status
                    )
                    compliance_profile.auditor_insights = (
                        financial_data["auditors_insights"]
                        or compliance_profile.auditor_insights
                    )
                    compliance_profile.pending_litigations = (
                        financial_data["pending_litigations"]
                        or compliance_profile.pending_litigations
                    )
                    compliance_profile.regulatory_filings = (
                        financial_data["regulatory_filings"]
                        or compliance_profile.regulatory_filings
                    )
                else:
                    compliance_profile = OrganizationComplianceProfiles(
                        organization_id=org_id,
                        compliance_status=financial_data["compliance_status"],
                        auditor_insights=financial_data["auditors_insights"],
                        pending_litigations=financial_data["pending_litigations"],
                        regulatory_filings=financial_data["regulatory_filings"],
                    )
                    db.session.add(compliance_profile)

                db.session.commit()
                flash("Financial overview saved successfully.", "success")
                return redirect(url_for("re.add_new_client", step=6))

            elif current_step == "6":
                # Step 7: Directors
                org_id = session.get("new_client_org_id")
                if not org_id:
                    flash("Please start from Step 1", "error")
                    return redirect(url_for("re.add_new_client"))

                directors_raw = request.form.get("directors_data", "[]")
                try:
                    directors_list = json.loads(directors_raw)
                except json.JSONDecodeError:
                    directors_list = []

                # Store in session
                session["step_6_data"] = {"directors": directors_list}

                for director_data in directors_list:
                    name = director_data.get("name", "").strip()
                    email = director_data.get("email", "").strip().lower()

                    if name and email:
                        new_director = OrganizationContacts(
                            organization_id=org_id,
                            contact_type="director",
                            name=name,
                            designation="Director",
                            email=email,
                            pancard=director_data.get("pan", "").strip(),
                            mobile=director_data.get("contact", "").strip(),
                        )
                        db.session.add(new_director)

                db.session.commit()

                org_name = session.get("new_client_org_name", "New Client")

                # Clear all session data
                session.pop("new_client_org_id", None)
                session.pop("new_client_org_name", None)
                session.pop("step_0_data", None)
                session.pop("step_1_data", None)
                session.pop("step_2_data", None)
                session.pop("step_3_data", None)
                session.pop("step_4_data", None)
                session.pop("step_5_data", None)
                session.pop("step_6_data", None)

                flash(f"New client '{org_name}' added successfully!", "success")
                return redirect(url_for("re.orgnization"))
        except Exception as e:
            db.session.rollback()
            import traceback

            print(f"Error saving data: {str(e)}")
            print(traceback.format_exc())
            flash(f"Error saving data: {str(e)}", "error")
            return redirect(url_for("re.add_new_client", step=current_step))

    # GET request - render the form with session data
    organization_types = OrganizationType.query.filter_by(active=True).all()
    constitution_types = Constitution.query.filter_by(active=True).all()
    departments = OrganizationDepartments.query.distinct(
        OrganizationDepartments.department_name
    ).all()

    step = request.args.get("step", 0, type=int)

    # Pass session data to template
    form_data = {
        "step_0": session.get("step_0_data", {}),
        "step_1": session.get("step_1_data", {}),
        "step_2": session.get("step_2_data", {}),
        "step_3": session.get("step_3_data", {}),
        "step_4": session.get("step_4_data", {}),
        "step_5": session.get("step_5_data", {}),
        "step_6": session.get("step_6_data", {}),
    }

    return render_template(
        "dashboards/re/profile.html",
        organization_types=organization_types,
        constitution_types=constitution_types,
        departments=departments,
        current_step=step,
        form_data=form_data,
    )


@re_bp.route("/profile")
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def profile():
    """
    Add New Client Page
    """
    try:
        # Fetch data needed for dropdowns
        organization_types = OrganizationType.query.filter_by(active=True).all()
        constitution_types = Constitution.query.filter_by(active=True).all()
        departments = OrganizationDepartments.query.distinct(
            OrganizationDepartments.department_name
        ).all()

        # Get current step from URL
        step = request.args.get("step", 0, type=int)

        # Pass session data to template (same as add_new_client route)
        form_data = {
            "step_0": session.get("step_0_data", {}),
            "step_1": session.get("step_1_data", {}),
            "step_2": session.get("step_2_data", {}),
            "step_3": session.get("step_3_data", {}),
            "step_4": session.get("step_4_data", {}),
            "step_5": session.get("step_5_data", {}),
            "step_6": session.get("step_6_data", {}),
        }

        return render_template(
            "dashboards/re/profile.html",
            organization_types=organization_types,
            constitution_types=constitution_types,
            departments=departments,
            current_step=step,
            form_data=form_data,  # Add this line
        )
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/training")
# @role_required()
def training():
    """
    Training Page
    """
    try:
        return render_template("training.html")
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/user", methods=["GET"])
# @role_required()
def user():
    """
    User Page
    """
    try:
        # Fetch all users from the database
        users = Users.query.all()
        return render_template("dashboards/re/user.html", users=users)
    except Exception as e:
        flash(f"Error fetching users: {str(e)}", "danger")
        return render_template("dashboards/re/user.html", users=[])


@re_bp.route("/links")
# @role_required()
def links():
    """
    Links Page
    """
    try:
        links = PolicyDocument.query.all()  # Fetch all links from the database
        print(links, links[0].url)
        return render_template("link.html", links=links)
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/profile/form", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def link_form():
    """
    Links Page
    """

    try:
        req = request.form
        print(req)
        org_name = request.form.get("org_name")
        legal_name = request.form.get("legal_name")
        constitution = request.form.get("constitution")
        # business_desc = request.form.get('business_desc')
        # no_branch_india = request.form.get('no_branch_india')
        # no_branch_out_india = request.form.get('no_branch_out_india')
        indian_regulatory = request.form.get("indian_regulatory")
        # international_regulatory = request.form.get('international_regulatory')
        # business_processes = request.form.get('business_processes')
        # org_history = request.form.get('org_history')
        # key_revenue_streams = request.form.get('key_revenue_streams')
        # key_markets = request.form.get('key_markets')
        # key_financial_metrics = request.form.get('key_financial_metrics')
        # last_year_total_rev = request.form.get('last_year_total_rev')
        # net_profit = request.form.get('net_profit')
        # total_assets = request.form.get('total_assets')
        # total_liabilities = request.form.get('total_liabilities')
        # key_financial_challenges = request.form.get('key_financial_challenges')
        # auditors_insight = request.form.get('auditors_insight')
        # compliance_status = request.form.get('compliance_status')
        # pending_litig = request.form.get('pending_litig')
        # reg_filing = request.form.get('reg_filing')
        # recent_significant = request.form.get('recent_significant')
        address = request.form.get("address")
        country = request.form.get("country")
        state = request.form.get("state")
        city = request.form.get("city")

        # try:
        #     departments = json.loads(request.form.get('departments', '[]'))
        # except json.JSONDecodeError:
        #     departments = []

        # try:
        #     org_structure_data = json.loads(request.form.get('org_structure', '[]'))
        # except json.JSONDecodeError:
        #     org_structure_data = []

        # Parse JSON fields
        try:
            industries = request.form.getlist("selected_industries")
        except json.JSONDecodeError:
            industries = {}

        try:
            organization_types = request.form.getlist("selected_org_types")
        except json.JSONDecodeError:
            organization_types = {}

        # try:
        #     head_office = json.loads(request.form.get('head_office', '{}'))
        # except json.JSONDecodeError:
        #     head_office = {}

        # try:
        #     key_locations = json.loads(request.form.get('key_locations', '[]'))
        # except json.JSONDecodeError:
        #     key_locations = []

        # try:
        #     directors = json.loads(request.form.get('directors', '[]'))
        # except json.JSONDecodeError:
        #     directors = []

        # print('org', industries)
        # Create Organization
        organization = Organizations(
            name=org_name,
            legal_name=legal_name,
            industry_type=industries,
            organization_type=organization_types,
            regulatory_status=indian_regulatory,
            constutution=constitution,
            incorporation_date=None,  # No corresponding field in the form
            fiscal_year_end=None,  # No corresponding field in the form
            status="active",
        )
        db.session.add(organization)
        db.session.flush()  # To get the organization_id

        # Handle head_office (JSON string)

        # Create Organization Address
        organization_address = OrganizationAddresses(
            organization_id=organization.organization_id,
            address_type="headoffice",  # Assuming this is the head office address
            address_line1=address,
            address_line2=None,  # No corresponding field in the form
            city=city,
            state=state,
            country=country,
            postal_code=None,  # No corresponding field in the form
            is_primary=True,  # Assuming the first address is primary
        )
        db.session.add(organization_address)

        # for idx, loc in enumerate(key_locations):
        #     organization_address = OrganizationAddresses(
        #         organization_id=organization.organization_id,
        #         address_type='keylocation',
        #         address_line1=loc.get('address'),
        #         address_line2=None,
        #         city=loc.get('city'),
        #         state=loc.get('state'),
        #         country=loc.get('country'),
        #         postal_code=None,
        #         is_primary=(idx == 0)  # Mark first address as primary
        #     )
        #     db.session.add(organization_address)

        # for dept_name in departments:
        #     if dept_name:  # Ensure it's not empty
        #         department = OrganizationDepartments(
        #             # organization_id=organization.organization_id,
        #             department_name=dept_name,
        #             # department_code=None,            # Optional: derive or pass if needed
        #             # parent_department_id=None,       # Optional: set if hierarchy is available
        #             # head_user_id=None                # Optional: set if known
        #         )
        #         db.session.add(department)

        # for item in org_structure_data:
        #     position = item.get('position')
        #     report_to = item.get('reportsTo')

        #     if position and report_to:  # Ensure required fields are present
        #         structure = OrganizationStructure(
        #             organization_id=organization.organization_id,
        #             position=position,
        #             report_to=report_to
        #         )
        #         db.session.add(structure)

        # # Handle directors (JSON string)
        # directors_str = req.get('directors')
        # directors = json.loads(directors_str) if directors_str else []

        # # Create Organization Contacts (Directors)
        # for director in directors:
        #     organization_contact = OrganizationContacts(
        #         organization_id=organization.organization_id,
        #         contact_type='director',
        #         name=director.get('name'),
        #         designation=None,  # No corresponding field in the form
        #         email=None,  # No corresponding field in the form
        #         pancard=director.get('pan'),
        #         mobile=director.get('contact'),
        #         is_active=True
        #     )
        #     db.session.add(organization_contact)

        # Create Organization Info
        # organization_info = OrganizationInfo(
        #     organization_id=organization.organization_id,
        #     business_desc=business_desc,
        #     history=org_history,
        #     key_events=recent_significant,
        #     key_revenue=key_revenue_streams,
        #     key_markets_customers=key_markets,
        #     key_financials=key_financial_metrics,
        #     total_revenue_last_year=last_year_total_rev,
        #     net_profit_loss=net_profit,
        #     total_assets=total_assets,
        #     total_liabilities=total_liabilities,
        #     key_financial_challenges=key_financial_challenges,
        #     is_active=True
        # )
        # db.session.add(organization_info)

        # # Create Organization Compliance Profile
        # organization_compliance_profile = OrganizationComplianceProfiles(
        #     organization_id=organization.organization_id,
        #     regulatory_body=req.get('indian_regulatory'),
        #     compliance_type=None,  # No corresponding field in the form
        #     business_process=business_processes,
        #     compliance_status=compliance_status,
        #     auditor_insights=auditors_insight,
        #     pending_litigations=pending_litig,
        #     regulatory_filings=reg_filing,
        #     indian_regulatory_compliance=indian_regulatory,
        #     international_regulatory_compliance=international_regulatory,
        #     last_audit_date=None,  # No corresponding field in the form
        #     next_audit_due=None,  # No corresponding field in the form
        #     risk_rating=None  # No corresponding field in the form
        # )
        # db.session.add(organization_compliance_profile)

        # Generate admin user details for the organization
        # admin_role_id = Roles.query.filter_by(name='ADMIN').first()
        # password = f"{org_name[:3].lower()}@1234"
        # email = f"admin@{org_name.lower()}.com"

        # Create the admin user
        # admin_user = Users(
        #     organization_id=organization.organization_id,
        #     email=email,
        #     phone_no="0000000000",  # Default phone number
        #     name="Admin",
        #     department_id=None,  # Set if applicable
        #     role_id=admin_role_id.role_id,
        #     user_type_id=None,  # Set if applicable
        #     password_hash=generate_password_hash(password),
        #     status="active"
        # )
        # db.session.add(admin_user)

        db.session.commit()
        if current_user:
            if current_user.auditor_profile_id:
                insert_stmt = auditor_client.insert().values(
                    audit_id=current_user.auditor_profile_id,
                    client_id=organization.organization_id,
                )
                db.session.execute(insert_stmt)
                db.session.commit()

        return redirect(url_for("re.orgnization"))

    except Exception as err:
        db.session.rollback()
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": str(err)}), 500


@re_bp.route("/check_user_progress")
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def check_user_progress():
    """
    Check user progress for the 3-step process
    """
    try:
        user_data = {
            "has_guidelines": False,
            "has_clients": False,
            "has_projects": False,
        }

        # Check if user has guidelines
        if current_user.auditor_profile_id:
            # Check guidelines
            guidelines_count = (
                db.session.query(Guidelines)
                .join(auditor_selected_guidelines)
                .filter(
                    auditor_selected_guidelines.c.audit_id
                    == current_user.auditor_profile_id
                )
                .count()
            )
            user_data["has_guidelines"] = guidelines_count > 0

            # Check clients (organizations linked to auditor)
            clients_count = (
                db.session.query(auditor_client)
                .filter(auditor_client.c.audit_id == current_user.auditor_profile_id)
                .count()
            )
            user_data["has_clients"] = clients_count > 0

            # Check projects
            projects_count = Projects.query.filter(
                Projects.auditing_firm == current_user.auditor_profile_id
            ).count()
            user_data["has_projects"] = projects_count > 0

        return jsonify(user_data)

    except Exception as err:
        current_app.logger.error(f"Error checking user progress: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/edit_re_profile", methods=["GET", "POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def edit_re_profile():
    org_id = request.args.get("org_id", type=int)
    if not org_id:
        flash("Organization ID is missing.", "error")
        return redirect(url_for("re.dashboard"))  # fallback route

    existing_locations = OrganizationAddresses.query.filter_by(
        organization_id=org_id, address_type="keylocation"
    ).all()

    organization = Organizations.query.filter_by(organization_id=org_id).first_or_404()

    # Fetch all organization types from the database
    organization_types = OrganizationType.query.filter_by(active=True).all()

    # Fetch all constitution types from the database
    constitution_types = Constitution.query.filter_by(active=True).all()

    # For edit mode, get the currently selected organization types
    selected_org_types = []
    if organization and organization.organization_type:
        selected_org_types = organization.organization_type

    # Fetch the head office address
    head_office = OrganizationAddresses.query.filter_by(
        organization_id=org_id, address_type="headoffice"
    ).first()

    # Fetch existing organization info and compliance profile data
    org_info = OrganizationInfo.query.filter_by(
        organization_id=org_id, is_active=True
    ).first()
    compliance_profile = OrganizationComplianceProfiles.query.filter_by(
        organization_id=org_id
    ).first()

    # Fetch existing organization structure
    existing_structures = OrganizationStructure.query.filter_by(
        organization_id=org_id
    ).all()

    # inside edit_re_profile
    existing_directors = OrganizationContacts.query.filter_by(
        organization_id=org_id, contact_type="director"
    ).all()

    print("this is existing directors", existing_directors)

    departments = OrganizationDepartments.query.distinct(
        OrganizationDepartments.department_name
    ).all()

    # Read association table to get saved department IDs (Option B)
    res = (
        db.session.query(department_in_org.c.department_id)
        .filter(department_in_org.c.organization_id == org_id)
        .all()
    )
    org_department_ids = [int(r[0]) for r in res]  # list of ints

    # For robust template comparisons you can also pass strings:
    org_department_ids_str = [str(x) for x in org_department_ids]

    if request.method == "POST":
        try:
            # Get the current step from the form
            current_step = request.form.get("current_step", "0")

            # Handle organization basic info
            org_name = request.form.get("org_name", "").strip()
            legal_name = request.form.get("legal_name", "").strip()
            constitution = request.form.get("constitution", "").strip()
            indian_regulatory = request.form.get("indian_regulatory", "").strip()

            # Handle organization types (checkboxes)

            selected_industries = request.form.getlist("selected_industries")

            # Handle organization types (checkboxes)
            selected_org_types = request.form.getlist("selected_org_types")

            # Update organization basic info
            if organization:
                organization.name = org_name or organization.name
                organization.legal_name = legal_name or organization.legal_name
                organization.constutution = constitution or organization.constutution
                organization.organization_type = (
                    selected_org_types  # This will be a list of selected types
                )
                organization.industry_type = (
                    selected_industries
                    if selected_industries
                    else organization.industry_type
                )

            # Handle head office address
            addr_line1 = request.form.get("address", "").strip()
            country = request.form.get("country", "").strip()
            state = request.form.get("state", "").strip()
            city = request.form.get("city", "").strip()

            if head_office:
                # Update existing head office
                head_office.address_line1 = addr_line1 or head_office.address_line1
                head_office.city = city or head_office.city
                head_office.state = state or head_office.state
                head_office.country = country or head_office.country
            else:
                # Create new head office entry
                head_office = OrganizationAddresses(
                    organization_id=org_id,
                    address_type="headoffice",
                    address_line1=addr_line1,
                    city=city,
                    state=state,
                    country=country,
                    is_primary=True,
                )
                db.session.add(head_office)

            db.session.commit()

            # If this is step 0 submission, stay on same page but move to next step
            if current_step == "0":
                flash(
                    "Organization and Head Office information updated successfully.",
                    "success",
                )
                return redirect(url_for("re.edit_re_profile", org_id=org_id, step=1))
                # Don't redirect, just reload the same template with updated data
                # The JavaScript will handle moving to the next step
            else:
                flash(
                    "Organization and Head Office information updated successfully.",
                    "success",
                )
                return redirect(url_for("re.edit_re_profile", org_id=org_id, step=1))

        except Exception as e:
            db.session.rollback()
            flash(f"Error saving organization: {str(e)}", "error")

    return render_template(
        "edit_re_org.html",
        organization=organization,
        org_info=org_info,  # Pass existing org info
        compliance_profile=compliance_profile,  # Pass existing compliance profile
        existing_structures=existing_structures,
        departments=departments,
        key_locations=existing_locations,
        org_department_ids=org_department_ids,
        org_department_ids_str=org_department_ids_str,
        existing_directors=existing_directors,
        head_office=head_office,
        organization_types=organization_types,  # Pass organization types to template
        selected_org_types=selected_org_types,  # Pass selected types
        constitution_types=constitution_types,  # Pass constitution types to template
    )


# route to handle editing of 'Other' key locations for an organization
@re_bp.route("/organization/<int:org_id>/edit-locations", methods=["GET", "POST"])
def edit_organization_locations(org_id):
    """
    Handle editing of 'Other' key locations for an organization.
    """
    # Fetch the organization to ensure it exists and user has permission
    organization = Organizations.query.get_or_404(org_id)

    # --- Add a security check here if needed, e.g. ---
    # if organization.owner_id != current_user.id:
    #     abort(403) # Forbidden

    if request.method == "POST":
        # Get the JSON string of locations from the hidden form input
        locations_data_str = request.form.get("locations_data")
        if not locations_data_str:
            flash("No location data submitted.", "warning")
            return redirect(
                url_for("main.edit_organization_locations", organization_id=org_id)
            )

        try:
            # Parse the JSON string into a Python list of dictionaries
            submitted_locations = json.loads(locations_data_str)
        except json.JSONDecodeError:
            flash("Invalid data format received. Please try again.", "danger")
            return redirect(
                url_for("main.edit_organization_locations", organization_id=org_id)
            )

        # --- Sync Database using the "Delete-Then-Add" Strategy ---
        # This is the simplest and most robust way to handle dynamic lists.

        # 1. Delete all existing 'Other' locations for this organization
        OrganizationAddresses.query.filter_by(
            organization_id=organization.organization_id,
            address_type="keylocation",  # Assuming the type for these is 'Other'
        ).delete()

        # 2. Add all the submitted locations as new records
        for loc in submitted_locations:
            if all(k in loc for k in ["address", "country", "state", "city"]):
                new_address = OrganizationAddresses(
                    organization_id=organization.organization_id,
                    address_type="keylocation",  # Set the type consistently
                    address_line1=loc["address"],
                    country=loc["country"],
                    state=loc["state"],
                    city=loc["city"],
                )
                db.session.add(new_address)

        # 3. Commit the transaction to the database
        db.session.commit()

        flash("Key locations have been updated successfully!", "success")
        # Redirect to the organization's profile page or another relevant page
        return redirect(url_for("re.edit_re_profile", org_id=org_id, step=2))
    # --- Handle GET Request ---
    # Fetch existing "Other" locations to populate the form on page load
    existing_locations = OrganizationAddresses.query.filter_by(
        organization_id=org_id, address_type="keylocation"
    ).all()

    return render_template(request.referrer)


# route to handle managing departments for an organization
@re_bp.route("/organization/<int:organization_id>/departments/manage", methods=["POST"])
def manage_departments(organization_id):
    organization = Organizations.query.get_or_404(organization_id)

    # Use getlist to read all checked checkbox values
    selected_ids = request.form.getlist("department_ids")  # ['1','3',...]
    # convert to ints and filter empty strings
    selected_department_ids = [int(x) for x in selected_ids if x and x.isdigit()]

    try:
        # Delete existing relationships
        delete_stmt = department_in_org.delete().where(
            department_in_org.c.organization_id == organization_id
        )
        db.session.execute(delete_stmt)

        # Optionally, insert new relationships using the association table
        if selected_department_ids:
            new_relationships = [
                {"organization_id": organization_id, "department_id": dept_id}
                for dept_id in selected_department_ids
            ]
            insert_stmt = department_in_org.insert().values(new_relationships)
            db.session.execute(insert_stmt)

        db.session.commit()
        flash("Departments updated successfully!", "success")
        return redirect(url_for("re.edit_re_profile", org_id=organization_id, step=3))
    except Exception as e:
        db.session.rollback()
        flash(f"An error occurred: {e}", "danger")
        return redirect(request.referrer)


# route to handle business overview form submission
@re_bp.route(
    "/organization/<int:org_id>/edit-business-overview", methods=["GET", "POST"]
)
def edit_organization_business_overview(org_id):
    """
    Handle editing of business overview information for an organization.
    """
    organization = Organizations.query.get_or_404(org_id)

    if request.method == "POST":
        try:
            # Get form data
            business_description = request.form.get("businessDescription", "").strip()
            branches_india = request.form.get("branchesIndia", type=int) or 0
            branches_outside_india = (
                request.form.get("branchesOutsideIndia", type=int) or 0
            )
            business_processes = request.form.get("businessProcesses", "").strip()
            org_history = request.form.get("orgHistory", "").strip()

            if not business_description:
                flash("Business description is required.", "error")
                return redirect(request.referrer)

            # Check if OrganizationInfo record exists
            org_info = OrganizationInfo.query.filter_by(
                organization_id=org_id, is_active=True
            ).first()

            if org_info:
                # Update existing record
                org_info.business_desc = business_description
                org_info.history = org_history
                org_info.no_of_branches_in_india = branches_india
                org_info.no_of_branches_outside_india = branches_outside_india
                org_info.updated_at = func.current_timestamp()
                flash("Business information updated successfully.", "success")
            else:
                # Create new record
                org_info = OrganizationInfo(
                    organization_id=org_id,
                    business_desc=business_description,
                    history=org_history or "",
                    key_events="",
                    key_revenue="",
                    key_markets_customers="",
                    key_financials="",
                    total_revenue_last_year="",
                    net_profit_loss="",
                    total_assets="",
                    total_liabilities="",
                    key_financial_challenges="",
                    no_of_branches_in_india=branches_india,
                    no_of_branches_outside_india=branches_outside_india,
                    is_active=True,
                )
                db.session.add(org_info)
                flash("Business information created successfully.", "success")

            # Handle business processes
            if business_processes:
                compliance_profile = OrganizationComplianceProfiles.query.filter_by(
                    organization_id=org_id
                ).first()

                if compliance_profile:
                    compliance_profile.business_process = business_processes
                    compliance_profile.updated_at = func.current_timestamp()
                else:
                    compliance_profile = OrganizationComplianceProfiles(
                        organization_id=org_id, business_process=business_processes
                    )
                    db.session.add(compliance_profile)

            # Commit changes
            db.session.commit()

            # IMPORTANT: Refresh the session to clear any cached data
            db.session.expunge_all()  # This removes all objects from session
            # OR use: db.session.refresh(org_info) if you want to refresh specific objects

            flash("Business overview has been updated successfully!", "success")
            return redirect(url_for("re.edit_re_profile", org_id=org_id, step=4))

        except Exception as e:
            db.session.rollback()
            flash(
                f"An error occurred while updating business information: {str(e)}",
                "error",
            )
            return redirect(request.referrer)

    # Handle GET Request - This shouldn't be reached normally since we redirect back
    return redirect(request.referrer)


@re_bp.route(
    "/organization/<int:org_id>/edit-organization-structure", methods=["GET", "POST"]
)
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def edit_organization_structure(org_id):
    """
    Handle editing of organization structure information for an organization.
    """
    # Fetch the organization to ensure it exists and user has permission
    organization = Organizations.query.get_or_404(org_id)

    # --- Add a security check here if needed, e.g. ---
    # if organization.owner_id != current_user.id:
    #     abort(403) # Forbidden

    if request.method == "POST":
        # Get the JSON string of positions from the hidden form input
        positions_data_str = request.form.get("positions_data")
        if not positions_data_str:
            flash("No organization structure data submitted.", "warning")
            return redirect(request.referrer)

        try:
            # Parse the JSON string into a Python list of dictionaries
            submitted_positions = json.loads(positions_data_str)
        except json.JSONDecodeError:
            flash("Invalid data format received. Please try again.", "danger")
            return redirect(request.referrer)

        # --- Sync Database using the "Delete-Then-Add" Strategy ---
        # This is the simplest and most robust way to handle dynamic lists.

        try:
            # 1. Delete all existing organization structure for this organization
            OrganizationStructure.query.filter_by(organization_id=org_id).delete()

            # 2. Add all the submitted positions as new records
            for position_data in submitted_positions:
                if all(k in position_data for k in ["position", "reportsTo"]):
                    new_structure = OrganizationStructure(
                        organization_id=org_id,
                        position=position_data["position"],
                        report_to=position_data["reportsTo"],
                    )
                    db.session.add(new_structure)

            # 3. Commit the transaction to the database
            db.session.commit()

            flash("Organization structure has been updated successfully!", "success")
            return redirect(url_for("re.edit_re_profile", org_id=org_id, step=5))

        except Exception as e:
            db.session.rollback()
            flash(
                f"An error occurred while updating organization structure: {str(e)}",
                "error",
            )
            return redirect(request.referrer)

    # --- Handle GET Request ---
    # Fetch existing organization structure to populate the form on page load
    existing_structures = OrganizationStructure.query.filter_by(
        organization_id=org_id
    ).all()

    # You can either render a specific template or redirect back to the main edit page
    # For now, redirecting back to referrer (similar to locations route)
    return redirect(re.orgnization)


# route to handle saving Financial Overview form
@re_bp.route("/organization/<int:organization_id>/financial-overview", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def save_financial_overview(organization_id):
    organization = Organizations.query.get_or_404(organization_id)

    # Read form fields
    key_revenue = request.form.get("revenueStreams", "").strip()
    key_markets_customers = request.form.get("marketsCustomers", "").strip()
    key_financials = request.form.get("financialMetrics", "").strip()
    total_revenue = request.form.get("totalRevenue", "").strip()
    net_profit_loss = request.form.get("netProfitLoss", "").strip()
    total_assets = request.form.get("totalAssets", "").strip()
    total_liabilities = request.form.get("totalLiabilities", "").strip()
    key_financial_challenges = request.form.get("financialChallenges", "").strip()

    auditors_insights = request.form.get("auditorsInsights", "").strip()
    compliance_status = request.form.get("complianceStatus", "").strip()
    pending_litigations = request.form.get("pendingLitigations", "").strip()
    regulatory_filings = request.form.get("regulatoryFilings", "").strip()

    try:
        # --- OrganizationInfo (create or update) ---
        org_info = OrganizationInfo.query.filter_by(
            organization_id=organization_id, is_active=True
        ).first()

        if not org_info:
            # OrganizationInfo has several non-nullable columns; initialize sensible defaults
            org_info = OrganizationInfo(
                organization_id=organization_id,
                business_desc="",
                history="",
                key_events="",
                key_revenue=key_revenue or "",
                key_markets_customers=key_markets_customers or "",
                key_financials=key_financials or "",
                total_revenue_last_year=total_revenue or "",
                net_profit_loss=net_profit_loss or "",
                total_assets=total_assets or "",
                total_liabilities=total_liabilities or "",
                key_financial_challenges=key_financial_challenges or "",
            )
            db.session.add(org_info)
        else:
            # update only the fields coming from this form
            org_info.key_revenue = key_revenue or org_info.key_revenue
            org_info.key_markets_customers = (
                key_markets_customers or org_info.key_markets_customers
            )
            org_info.key_financials = key_financials or org_info.key_financials
            org_info.total_revenue_last_year = (
                total_revenue or org_info.total_revenue_last_year
            )
            org_info.net_profit_loss = net_profit_loss or org_info.net_profit_loss
            org_info.total_assets = total_assets or org_info.total_assets
            org_info.total_liabilities = total_liabilities or org_info.total_liabilities
            org_info.key_financial_challenges = (
                key_financial_challenges or org_info.key_financial_challenges
            )

        # --- OrganizationComplianceProfiles (create or update) ---
        compliance_profile = OrganizationComplianceProfiles.query.filter_by(
            organization_id=organization_id
        ).first()

        if not compliance_profile:
            compliance_profile = OrganizationComplianceProfiles(
                organization_id=organization_id,
                compliance_status=compliance_status or "",
                auditor_insights=auditors_insights or "",
                pending_litigations=pending_litigations or "",
                regulatory_filings=regulatory_filings or "",
            )
            db.session.add(compliance_profile)
        else:
            compliance_profile.compliance_status = (
                compliance_status or compliance_profile.compliance_status
            )
            compliance_profile.auditor_insights = (
                auditors_insights or compliance_profile.auditor_insights
            )
            compliance_profile.pending_litigations = (
                pending_litigations or compliance_profile.pending_litigations
            )
            compliance_profile.regulatory_filings = (
                regulatory_filings or compliance_profile.regulatory_filings
            )

        db.session.commit()
        flash("Financial Overview saved successfully.", "success")
        # Redirect back to the edit form (preserve org_id in querystring)
        return redirect(url_for("re.edit_re_profile", org_id=organization_id, step=6))
    except Exception as e:
        db.session.rollback()
        # current_app.logger.exception("Error saving financial overview for org %s", organization_id)
        flash(f"An error occurred while saving Financial Overview: {e}", "danger")
        return redirect(request.referrer)


# route to handle saving Directors form


@re_bp.route("/organization/<int:organization_id>/directors", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def save_directors(organization_id):
    organization = Organizations.query.get_or_404(organization_id)

    # Read the JSON payload placed into the hidden input 'directors_data'
    directors_raw = request.form.get("directors_data", "").strip()

    print(f"DEBUG: Received directors_raw: {directors_raw}")
    print(f"DEBUG: Organization ID: {organization_id}")

    try:
        if not directors_raw:
            flash("No director data received.", "warning")
            print("DEBUG: No director data received")
            return redirect(url_for("re.orgnization"))  # Redirect to organization page

        directors_list = json.loads(directors_raw)
        print(f"DEBUG: Parsed directors_list: {directors_list}")

        if not isinstance(directors_list, list):
            raise ValueError("directors_data must be a JSON list")

        # normalize incoming emails and build a lookup
        incoming_emails = set()
        incoming_by_email = {}
        for d in directors_list:
            email = (d.get("email") or "").strip().lower()
            if not email:
                continue
            incoming_emails.add(email)
            incoming_by_email[email] = {
                "name": d.get("name", "").strip(),
                "pancard": d.get("pan", "").strip(),
                "mobile": d.get("contact", "").strip(),
            }

        print(f"DEBUG: Processed directors: {incoming_by_email}")

        # fetch ALL director contacts for this organization
        existing_directors = OrganizationContacts.query.filter_by(
            organization_id=organization_id, contact_type="director"
        ).all()
        existing_by_email = {
            (c.email or "").strip().lower(): c for c in existing_directors if c.email
        }

        print(f"DEBUG: Found {len(existing_directors)} existing directors")

        # update or create
        for email, payload in incoming_by_email.items():
            if email in existing_by_email:
                c = existing_by_email[email]
                c.name = payload["name"] or c.name
                c.pancard = payload["pancard"] or c.pancard
                c.mobile = payload["mobile"] or c.mobile
                print(f"DEBUG: Updated director: {email}")
            else:
                new_c = OrganizationContacts(
                    organization_id=organization_id,
                    contact_type="director",
                    name=payload["name"] or "Director",
                    designation="Director",
                    email=email,
                    pancard=payload["pancard"] or None,
                    mobile=payload["mobile"] or None,
                )
                db.session.add(new_c)
                print(f"DEBUG: Created new director: {email}")

        # Delete directors that are no longer in the list
        directors_to_delete = []
        for email, c in existing_by_email.items():
            if email not in incoming_emails:
                directors_to_delete.append(c)
                print(f"DEBUG: Marked for deletion: {email}")

        # Delete the directors
        for director in directors_to_delete:
            db.session.delete(director)
            print(f"DEBUG: Deleted director: {director.email}")

        db.session.commit()
        print("DEBUG: Successfully committed directors to database")
        flash("Directors saved successfully.", "success")
        return redirect(
            url_for("re.orgnization")
        )  # Redirect to organization page after success

    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Error saving directors: {str(e)}")
        current_app.logger.exception(
            "Error saving directors for org %s", organization_id
        )
        flash(f"An error occurred while saving directors: {e}", "danger")
        return redirect(
            url_for("re.orgnization")
        )  # Redirect to organization page even on error


@re_bp.route("/user/add", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def add_user():
    AUDITOR_ROLE_NAME = "AUDITOR"
    RE_ROLE_NAME = "RE"

    try:
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")
        email = request.form.get("email")
        phone_no = request.form.get("phone_no")
        role_id_str = request.form.get("role")
        status = request.form.get("status", "active")
        password = request.form.get("password")

        if not all([first_name, last_name, email, phone_no, role_id_str, password]):
            flash(
                "All required fields (Name, Email, Phone, Role, Password) must be provided.",
                "danger",
            )
            current_app.logger.warning("Missing required form data for user creation.")
            return redirect(request.referrer)

        if "@" not in email or "." not in email:
            flash("Invalid email format.", "danger")
            current_app.logger.warning(f"Invalid email format provided: {email}")
            return redirect(request.referrer)

        if len(password) < 8:
            flash("Password must be at least 8 characters long.", "danger")
            current_app.logger.warning("Password too short.")
            return redirect(request.referrer)

        try:
            role_id = int(role_id_str)
        except ValueError:
            flash("Invalid role ID provided.", "danger")
            current_app.logger.error(f"Invalid role ID format: {role_id_str}")
            return redirect(request.referrer)

        full_name = f"{first_name.capitalize()} {last_name.capitalize()}"

        existing_user = db.session.execute(
            select(Users).filter_by(email=email)
        ).scalar_one_or_none()
        if existing_user:
            flash(f'User with email "{email}" already exists.', "danger")
            current_app.logger.warning(
                f"Attempted to create duplicate user with email: {email}"
            )
            return redirect(request.referrer)

        role_obj = Roles.query.filter_by(role_id=role_id).first()
        if not role_obj:
            flash("Selected role does not exist.", "danger")
            current_app.logger.error(f"Role with ID {role_id} not found in database.")
            return redirect(request.referrer)

        role_name = role_obj.name

        organization_id_for_user = None
        auditor_profile_id_for_user = None

        if role_name == AUDITOR_ROLE_NAME:
            if current_user.auditor_profile_id is None:
                flash(
                    "Auditor profile ID not set for this context. Cannot assign Auditor role.",
                    "danger",
                )
                current_app.logger.error(
                    "current_app.auditor_profile_id is missing when trying to assign AUDITOR role."
                )
                return redirect(request.referrer)
            auditor_profile_id_for_user = current_user.auditor_profile_id
            current_app.logger.info(
                f"Assigning AUDITOR role with auditor_profile_id: {auditor_profile_id_for_user}"
            )

        elif role_name == RE_ROLE_NAME:
            if current_user.organization_id is None:
                flash(
                    "Organization ID not set for this context. Cannot assign RE role.",
                    "danger",
                )
                current_app.logger.error(
                    "current_app.organization_id is missing when trying to assign RE role."
                )
                return redirect(request.referrer)
            organization_id_for_user = current_user.organization_id
            current_app.logger.info(
                f"Assigning RE role with organization_id: {organization_id_for_user}"
            )

        else:
            flash(
                f'The role "{role_name}" cannot be assigned to new users through this process.',
                "danger",
            )
            current_app.logger.warning(
                f"Attempted to assign unassignable role: {role_name}"
            )
            return redirect(request.referrer)

        password_hash = generate_password_hash(password)

        new_user = Users(
            name=full_name,
            email=email,
            phone_no=phone_no,
            role_id=role_id,
            organization_id=organization_id_for_user,
            auditor_profile_id=auditor_profile_id_for_user,
            password_hash=password_hash,
            status=status,
        )
        current_app.logger.info(f"New user object: {new_user}")

        db.session.add(new_user)
        db.session.flush()
        current_app.logger.info("User added to session, flushing successful.")

        db.session.commit()
        current_app.logger.info("User committed to the database.")

        flash("User added successfully!", "success")
        return redirect(url_for("re.admin"))

    except IntegrityError as e:
        db.session.rollback()
        error_message = (
            "A database integrity error occurred (e.g., duplicate unique entry)."
        )
        current_app.logger.error(f"{error_message}: {str(e)}")
        flash(f"Error adding user: {error_message}", "danger")
        return redirect(request.referrer)
    except SQLAlchemyError as e:
        db.session.rollback()
        error_message = "A database error occurred during user creation."
        current_app.logger.error(f"{error_message}: {str(e)}")
        flash(f"Error adding user: {error_message}", "danger")
        return redirect(request.referrer)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(
            f"An unexpected error occurred while adding user: {str(e)}", exc_info=True
        )
        flash(f"An unexpected error occurred: {str(e)}", "danger")
        return redirect(request.referrer)


@re_bp.route("/add-user-permissions", methods=["GET", "POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def add_user_permissions():
    """
    Endpoint to add a user role and permissions.
    """
    if request.method == "POST":
        try:
            # Extract form data
            name = request.form.get("name")
            description = request.form.get("description")
            permissions = request.form.getlist(
                "permissions"
            )  # Get permissions as a list

            # Convert permissions list to JSON format
            permissions_json = {name: permissions}

            # Create a new role
            new_role = Roles(
                name=name, description=description, permissions=permissions_json
            )
            db.session.add(new_role)
            db.session.commit()

            flash("Role added successfully!", "success")
            return redirect(url_for("re.add_user_permissions"))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error adding role: {str(e)}")
            flash("An error occurred while adding the role.", "danger")
            return redirect(url_for("re.add_user_permissions"))

    return redirect(url_for("re.list_roles"))


@re_bp.route("/roles", methods=["GET"])
# # @role_required()
def list_roles():
    """
    Endpoint to list all roles.
    """
    try:

        roles = Roles.query.all()
        routes = list_routes(create_app())
        url = [route["url"] for route in routes]
        print(url)
        return render_template("add_user.html", roles=roles, urls=url)
    except Exception as e:
        current_app.logger.error(f"Error fetching roles: {str(e)}")
        flash("An error occurred while fetching roles.", "danger")
        return redirect(url_for("re.dashboard"))


@re_bp.route("/roles/edit/<int:role_id>", methods=["GET", "POST"])
# # @role_required()
def edit_role(role_id):
    """
    Endpoint to edit a role.
    """
    role = Roles.query.get_or_404(role_id)
    routes = list_routes(create_app())
    url = [route["url"] for route in routes]
    print("urls", url)
    selected_permissions = role.permissions.get(role.name, [])
    if request.method == "POST":

        try:
            role.name = request.form.get("name")
            role.description = request.form.get("description")
            permissions = request.form.getlist("permissions")
            print(permissions)
            role.permissions = {role.name: permissions}
            db.session.commit()
            flash("Role updated successfully!", "success")
            return redirect(url_for("re.list_roles"))
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating role: {str(e)}")
            flash("An error occurred while updating the role.", "danger")
            return redirect(url_for("re.edit_role", role_id=role_id))
    return render_template(
        "add_user.html", role=role, urls=url, selected_permissions=selected_permissions
    )


@re_bp.route("/roles/delete/<int:role_id>", methods=["POST"])
# # @role_required()
def delete_role(role_id):
    """
    Endpoint to delete a role.
    """
    try:
        role = Roles.query.get_or_404(role_id)
        if role:
            db.session.delete(role)
            db.session.commit()
            flash("Role deleted successfully!", "success")
        else:
            flash("Role not found.", "warning")
        return redirect(url_for("re.list_roles"))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting role: {str(e)}")
        flash("An error occurred while deleting the role.", "danger")
        return redirect(url_for("re.list_roles"))


@re_bp.route("/clause", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def clause():
    """
    Clause Page
    """
    add_to_breadcrumb(request.full_path, "Clause")
    try:
        data = request.get_json()
        guideline_id = data.get("id")

        if not guideline_id:
            raise ValueError("Guideline ID is required")

        if guideline_id:
            session["clause_id"] = guideline_id
            clauses = Clauses.query.filter_by(guideline_id=guideline_id).all()
        else:
            clauses = Clauses.query.all()

        # Prepare data for rendering
        clause_data = [
            {
                "id": clause.id,
                "clause_no": clause.clause_no,
                "clause_text": clause.clause_text,
                "compliance_activities": clause.compliance_activities,
            }
            for clause in clauses
        ]

        # fetch guideline and consolidated evidence (if any)
        guideline = Guidelines.query.get(guideline_id)
        guideline_name = None
        if guideline and guideline.guideline_data:
            guideline_name = guideline.guideline_data.get("DocumentDetails", {}).get(
                "DocumentName"
            )

        consolidated_evidence = None
        consolidated_rec = ComplifyreConsolidatedEvidence.query.filter_by(
            guideline_id=guideline_id
        ).first()

        # FIX: Properly handle consolidated_evidence data type
        if consolidated_rec and consolidated_rec.consolidate_evidence:
            if isinstance(consolidated_rec.consolidate_evidence, str):
                try:
                    consolidated_evidence = json.loads(
                        consolidated_rec.consolidate_evidence
                    )
                except json.JSONDecodeError:
                    current_app.logger.error(
                        f"Failed to parse consolidate_evidence as JSON for guideline {guideline_id}"
                    )
                    consolidated_evidence = None
            else:
                consolidated_evidence = consolidated_rec.consolidate_evidence

        # Calculate pending activities count
        total_clauses = len(clauses)
        clauses_with_activities = sum(
            1 for clause in clauses
            if clause.compliance_activities and len(clause.compliance_activities) > 0
        )
        pending_clauses = total_clauses - clauses_with_activities
        total_activities = sum(
            len(clause.compliance_activities) for clause in clauses
            if clause.compliance_activities
        )

        return render_template(
            "clause.html",
            clauses=clause_data,
            guideline_id=guideline_id,
            guideline=guideline,
            guideline_name=guideline_name,
            consolidated_evidence=consolidated_evidence,
            total_clauses=total_clauses,
            clauses_with_activities=clauses_with_activities,
            pending_clauses=pending_clauses,
            total_activities=total_activities,
        )
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": f"Internal server error {err}"}), 500


@re_bp.route("/clause", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def get_clause():
    add_to_breadcrumb(request.full_path, "All Clause")
    try:
        id_param = request.args.get("guideline_id")
        guideline = None
        guideline_name = None
        guideline_id = None

        if id_param:
            guideline_id = int(id_param)
            guideline = Guidelines.query.get(guideline_id)
            if guideline and guideline.guideline_data:
                guideline_name = guideline.guideline_data.get(
                    "DocumentDetails", {}
                ).get("DocumentName")
            clauses = Clauses.query.filter_by(guideline_id=guideline_id).all()
        else:
            clauses = Clauses.query.all()

        # Natural sorting function (same as in your extract_clauses task)
        def natural_sort_key(clause):
            text = clause.clause_no
            if text is None or text == "":
                return [float("inf")]  # Put empty/missing numbers at the end
            return [
                int(part) if part.isdigit() else part.lower()
                for part in re.split(r"(\d+)", str(text))
            ]

        # Sort clauses using natural sorting
        sorted_clauses = sorted(clauses, key=natural_sort_key)
        clause_data = sorted_clauses

        consolidated_evidence = None
        if guideline_id:
            consolidated_rec = ComplifyreConsolidatedEvidence.query.filter_by(
                guideline_id=guideline_id
            ).first()
            if consolidated_rec and consolidated_rec.consolidate_evidence:
                if isinstance(consolidated_rec.consolidate_evidence, str):
                    try:
                        consolidated_evidence = json.loads(
                            consolidated_rec.consolidate_evidence
                        )
                    except json.JSONDecodeError:
                        current_app.logger.error(
                            f"Failed to parse consolidate_evidence as JSON for guideline {guideline_id}"
                        )
                        consolidated_evidence = {"grouped_evidences": []}
                else:
                    consolidated_evidence = consolidated_rec.consolidate_evidence

        return render_template(
            "clause.html",
            clause_data=clause_data,
            guideline_id=guideline_id,
            guideline=guideline,
            guideline_name=guideline_name,
            consolidated_evidence=consolidated_evidence,
            now=datetime.now(),
        )
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}", exc_info=True)
        return jsonify({"error": f"Internal server error {err}"}), 500


@re_bp.route("/clauses/<clause_id>/number", methods=["PUT"])
def update_clause_number(clause_id):
    print(f"DEBUG: Route hit with clause_id: {clause_id}")
    try:
        if not request.is_json:
            return (
                jsonify(
                    {"success": False, "error": "Content-Type must be application/json"}
                ),
                400,
            )

        data = request.get_json()
        new_clause_number = data.get("clause_number")

        if not new_clause_number:
            return (
                jsonify({"success": False, "error": "Clause number is required"}),
                400,
            )

        # Get the clause
        clause = Clauses.query.get(clause_id)
        if not clause:
            return jsonify({"success": False, "error": "Clause not found"}), 404

        print(f"DEBUG: Before update - clause_no: '{clause.clause_no}'")

        # Update using direct assignment
        clause.clause_no = new_clause_number.strip()

        # Force the session to see the change
        db.session.add(clause)

        # Commit the transaction
        db.session.commit()

        # Refresh to get the updated value from database
        db.session.refresh(clause)
        print(f"DEBUG: After commit - clause_no: '{clause.clause_no}'")

        return jsonify(
            {
                "success": True,
                "message": "Clause number updated successfully",
                "clause_number": clause.clause_no,
            }
        )

    except Exception as e:
        db.session.rollback()
        print(f"DEBUG: Error: {str(e)}")
        import traceback

        print(f"DEBUG: Traceback: {traceback.format_exc()}")
        return jsonify({"success": False, "error": str(e)}), 500


@re_bp.route("/compliance_activities", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def compliance_activities():
    """
    Compliance Activities Page
    """
    try:
        pdf_service = PDFService()
        data = request.get_json()
        id = data.get("id")
        clauses = Clauses.query.filter_by(id=id).first()

        if not clauses:
            return jsonify({"error": "Clause not found"}), 404

        guideline_id = clauses.guideline_id

        if not guideline_id:
            raise ValueError("Guideline ID is required")

        guideline = Guidelines.query.filter_by(id=guideline_id).first()
        if not guideline:
            raise ValueError("Guideline not found")

        file_record = File.query.filter_by(id=guideline.file_id).first()
        if not file_record:
            raise ValueError("File record not found for guideline")

        vec_id = file_record.vector_store_id

        file_url = File.query.filter_by(id=guideline.file_id).first()
        url = file_url.path

        text = pdf_service.extract_text_from_pdf(url)
        compliance_data = pdf_service.retrive_regulatory_complience(
            clauses.clause_text, text
        )
        claus_json = json.loads(f"""{compliance_data}""")

        compliance_activities = []
        comps_to_process = []  # Store activities that need test procedures

        # Fix: Proper loop structure - iterate over claus_json["compliance_activities"]
        for index, item in enumerate(claus_json["compliance_activities"], start=1):
            print(item)

            # Extract compliance_level from the AI response, default to "Design" if not provided
            compliance_level = item.get("compliance_level", "Design")

            # Validate compliance_level value
            valid_compliance_levels = [
                "Design",
                "Implementation",
                "Operating Effectiveness",
            ]
            if compliance_level not in valid_compliance_levels:
                compliance_level = "Design"  # Default to Design if invalid

            # Ensure activity_id is numerical - use AI response if numerical, otherwise use index
            activity_id_from_ai = item.get("activity_id", "")

            # Try to extract numerical value, if fails use the index
            try:
                # Extract numbers from string if it contains numbers
                import re

                numbers = re.findall(r"\d+", str(activity_id_from_ai))
                if numbers:
                    activity_id = numbers[0]  # Take first number found
                else:
                    activity_id = str(index)  # Use the loop index as fallback
            except (ValueError, TypeError):
                activity_id = str(index)  # Use the loop index as fallback

            comp = ComplianceActivities(
                clause_id=id,
                relevant_departments_id=int(item["department_id"]),
                relevant_departments=item["relevant_departments"],
                process=item["process_name"],
                sub_process=item["sub_process_name"],
                activity_id=activity_id,  # Use the validated activity_id instead of item["activity_id"]
                activity_description=item["activity_description"],
                responsible_party=item["responsible_party"],
                frequency=item["frequency"],
                evidence_required=item["evidence_required"],
                compliance_level=compliance_level,
            )
            db.session.add(comp)
            db.session.flush()  # Flush to get the comp.id

            compliance_activities.append(comp)

            # Store for test procedure processing
            comps_to_process.append(
                (comp.id, clauses.clause_text, json.dumps(item), vec_id)
            )

        # Add to the database
        db.session.commit()

        # Step 2: Generate test procedures for each compliance activity
        for (
            comp_id_val,
            clause_text_val,
            comp_payload_json,
            vec_id_val,
        ) in comps_to_process:
            try:
                process_test_procedures(
                    comp_id=comp_id_val,
                    clause_text=clause_text_val,
                    compliance_activity_payload=comp_payload_json,
                    vec_id=vec_id_val,
                )
                current_app.logger.info(
                    f"Test procedures generated for compliance activity {comp_id_val}"
                )
            except Exception as te:
                current_app.logger.error(
                    f"Error generating test procedures for comp(id={comp_id_val}): {te}"
                )

        current_app.logger.info(
            "Compliance activities and test procedures committed to the database."
        )
        return redirect(request.referrer)

    except Exception as err:
        db.session.rollback()
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/regenerate_compliance_activities", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def regenerate_compliance_activities():
    """
    Regenerate compliance activities by deleting existing ones and creating new ones
    """
    try:
        pdf_service = PDFService()
        data = request.get_json()
        id = data.get("id")
        clauses = Clauses.query.filter_by(id=id).first()

        if not clauses:
            return jsonify({"error": "Clause not found"}), 404

        guideline_id = clauses.guideline_id

        if not guideline_id:
            raise ValueError("Guideline ID is required")

        guideline = Guidelines.query.filter_by(id=guideline_id).first()
        if not guideline:
            raise ValueError("Guideline not found")

        file_record = File.query.filter_by(id=guideline.file_id).first()
        if not file_record:
            raise ValueError("File record not found for guideline")

        vec_id = file_record.vector_store_id

        # Delete existing compliance activities and their related records
        existing_activities = ComplianceActivities.query.filter_by(clause_id=id).all()

        for activity in existing_activities:
            # Delete related records first (in correct order to respect foreign key constraints)

            # 1. First delete project_compliance_activities records
            project_activities = ProjectComplianceActivity.query.filter_by(
                original_activity_id=activity.id
            ).all()
            for project_activity in project_activities:
                db.session.delete(project_activity)

            # 2. Delete test_procedures records (using activity_id column)
            test_procedures = TestProcedures.query.filter_by(
                activity_id=activity.id
            ).all()
            for test_procedure in test_procedures:
                db.session.delete(test_procedure)

            # 3. Delete control_activities records (delete test_steps first)
            control_activities = ControlActivity.query.filter_by(
                compliance_activity_id=activity.id
            ).all()
            for control_activity in control_activities:
                # Delete test_steps first (foreign key constraint)
                from app.models.ai import TestSteps
                test_steps = TestSteps.query.filter_by(
                    control_id=control_activity.id
                ).all()
                for ts in test_steps:
                    db.session.delete(ts)
                db.session.flush()
                db.session.delete(control_activity)

            # 4. Delete how_to_perform_activity records (using activity_id column)
            how_to_perform = HowToPerformActivity.query.filter_by(
                activity_id=activity.id
            ).all()
            for how_to in how_to_perform:
                db.session.delete(how_to)

            # 5. Skip projects deletion — projects should not be deleted on activity regeneration

            # Finally delete the compliance activity itself
            db.session.delete(activity)

        db.session.commit()

        # Now generate new activities
        file_url = File.query.filter_by(id=guideline.file_id).first()
        url = file_url.path

        text = pdf_service.extract_text_from_pdf(url)
        compliance_data = pdf_service.retrive_regulatory_complience(
            clauses.clause_text, text
        )
        claus_json = json.loads(f"""{compliance_data}""")

        compliance_activities = []
        comps_to_process = []  # Store activities that need test procedures

        # Fix: Proper loop structure - iterate over claus_json["compliance_activities"]
        for index, item in enumerate(claus_json["compliance_activities"], start=1):
            # Extract compliance_level with validation
            compliance_level = item.get("compliance_level", "Design")
            valid_compliance_levels = [
                "Design",
                "Implementation",
                "Operating Effectiveness",
            ]
            if compliance_level not in valid_compliance_levels:
                compliance_level = "Design"

            # Ensure activity_id is numerical - use AI response if numerical, otherwise use index
            activity_id_from_ai = item.get("activity_id", "")

            # Try to extract numerical value, if fails use the index
            try:
                # Extract numbers from string if it contains numbers
                import re

                numbers = re.findall(r"\d+", str(activity_id_from_ai))
                if numbers:
                    activity_id = numbers[0]  # Take first number found
                else:
                    activity_id = str(index)  # Use the loop index as fallback
            except (ValueError, TypeError):
                activity_id = str(index)  # Use the loop index as fallback

            comp = ComplianceActivities(
                clause_id=id,
                relevant_departments_id=int(item["department_id"]),
                relevant_departments=item["relevant_departments"],
                process=item["process_name"],
                sub_process=item["sub_process_name"],
                activity_id=activity_id,  # Use the validated activity_id
                activity_description=item["activity_description"],
                responsible_party=item["responsible_party"],
                frequency=item["frequency"],
                evidence_required=item["evidence_required"],
                compliance_level=compliance_level,
            )
            db.session.add(comp)
            db.session.flush()  # Flush to get the comp.id

            compliance_activities.append(comp)

            # Store for test procedure processing
            comps_to_process.append(
                (comp.id, clauses.clause_text, json.dumps(item), vec_id)
            )

        db.session.commit()

        # Step 2: Generate test procedures for each compliance activity
        for (
            comp_id_val,
            clause_text_val,
            comp_payload_json,
            vec_id_val,
        ) in comps_to_process:
            try:
                process_test_procedures(
                    comp_id=comp_id_val,
                    clause_text=clause_text_val,
                    compliance_activity_payload=comp_payload_json,
                    vec_id=vec_id_val,
                )
                current_app.logger.info(
                    f"Test procedures generated for compliance activity {comp_id_val}"
                )
            except Exception as te:
                current_app.logger.error(
                    f"Error generating test procedures for comp(id={comp_id_val}): {te}"
                )

        current_app.logger.info(
            f"Compliance activities and test procedures regenerated for clause_id={id}"
        )
        return (
            jsonify(
                {"message": "Activities and test procedures regenerated successfully"}
            ),
            200,
        )

    except Exception as err:
        db.session.rollback()
        current_app.logger.error(f"Error regenerating activities: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/generate_all_compliance", methods=["POST"])
# @role_required()
def generate_all_compliance():
    try:
        pdf_service = PDFService()
        data = request.get_json()
        clause_ids = data.get("clause_ids", [])
        total_success = 0
        total_failed = 0

        for id in clause_ids:
            try:
                clauses = Clauses.query.filter_by(id=id).first()
                if not clauses:
                    total_failed += 1
                    continue

                guideline_id = clauses.guideline_id
                if not guideline_id:
                    total_failed += 1
                    continue

                guideline = Guidelines.query.filter_by(id=guideline_id).first()
                file_url = File.query.filter_by(id=guideline.file_id).first()

                if not file_url:
                    total_failed += 1
                    continue

                text = pdf_service.extract_text_from_pdf(file_url.path)
                compliance_data = pdf_service.retrive_regulatory_complience(
                    clauses.clause_text, text
                )
                claus_json = json.loads(compliance_data)

                compliance_activities = []
                for item in claus_json["compliance_activities"]:
                    comp = ComplianceActivities(
                        clause_id=id,
                        relevant_departments_id=int(item["department_id"]),
                        relevant_departments=item["relevant_departments"],
                        process=item["process_name"],
                        sub_process=item["sub_process_name"],
                        activity_id=item["activity_id"],
                        activity_description=item["activity_description"],
                        responsible_party=item["responsible_party"],
                        frequency=item["frequency"],
                        evidence_required=item["evidence_required"],
                    )
                    compliance_activities.append(comp)

                db.session.add_all(compliance_activities)
                db.session.flush()
                total_success += 1
            except Exception as inner_err:
                current_app.logger.error(
                    f"Error processing clause {id}: {str(inner_err)}"
                )
                total_failed += 1

        db.session.commit()
        return (
            jsonify(
                {
                    "message": f"Compliance generation complete. Success: {total_success}, Failed: {total_failed}"
                }
            ),
            200,
        )

    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/activity_mapping", methods=["POST"])
# @role_required()
def activity_mapping_redundancy_check():
    """
    Processes a clause, extracts compliance data, and stores it in the database.
    """
    try:
        pdf_service = PDFService()
        data = request.get_json()
        id = data.get("id")
        print(id)
        clauses = Clauses.query.filter_by(id=id).first()
        print(clauses, clauses.clause_text)
        guideline_id = clauses.guideline_id
        print(guideline_id)
        if not guideline_id:
            raise ValueError("Guideline ID is required")
        # Query the database for the guideline and its associated clauses
        guideline = Guidelines.query.filter_by(id=guideline_id).first()
        if not guideline:
            raise ValueError("Guideline not found")
        file_url = File.query.filter_by(id=guideline.file_id).first()
        print(file_url)
        # Prepare data for rendering
        url = file_url.path
        print(url)
        text = pdf_service.extract_text_from_pdf(url)
        json_data = pdf_service.retrive_activity(clauses.clause_text, text)
        compliance_data = json.loads(f"""{json_data}""")
        # Create ActivityGuideline
        guideline = ActivityGuideline(
            activity_code=compliance_data.get("activity_code"),
            activity=compliance_data.get("activity"),
            redundancy_check=compliance_data.get("redundancy_check"),
            risk_level=compliance_data.get("risk_level"),
            mitigation_actions=compliance_data.get("mitigation_actions"),
            frequency=compliance_data.get("frequency"),
        )

        # Handle relationships
        # Key Owner
        key_owner_name = compliance_data.get("relevant_departments", {}).get(
            "key_owner"
        )
        if key_owner_name:
            key_owner = OrganizationDepartments.query.filter_by(
                name=key_owner_name
            ).first()
            if key_owner:
                guideline.key_owner = key_owner

        # Supporting Teams
        supporting_team_names = compliance_data.get("relevant_departments", {}).get(
            "supporting_teams", []
        )
        for team_name in supporting_team_names:
            team = OrganizationDepartments.query.filter_by(name=team_name).first()
            if team:
                guideline.supporting_teams.append(team)

        # Impacted Departments
        impacted_processes = compliance_data.get("impacted_processes_sub_processes", [])
        for process_info in impacted_processes:
            department_name = process_info.get("process")
            if department_name:
                department = OrganizationDepartments.query.filter_by(
                    name=department_name
                ).first()
                if department:
                    guideline.impacted_departments.append(department)

        # ClauseIntentAnalysis
        intent_analysis_data = compliance_data.get("clause_intent_analysis", {})
        intent_analysis = ClauseIntentAnalysis(
            intent=intent_analysis_data.get("intent"),
            regulatory_expectations=intent_analysis_data.get("regulatory_expectations"),
            risk_areas=intent_analysis_data.get("risk_areas"),
            operational_impact=intent_analysis_data.get("operational_impact"),
            core_purpose=intent_analysis_data.get("core_purpose"),
        )
        guideline.intent_analysis = intent_analysis

        # HowToPerform
        how_to_data = compliance_data.get("how_to_perform", {})
        how_to_perform = HowToPerform(
            execution_steps=how_to_data.get("execution_steps"),
            responsible_roles=how_to_data.get("responsible_roles"),
            timelines=how_to_data.get("timelines"),
        )
        guideline.how_to_perform = how_to_perform

        # EvidencesArtifacts
        evidence_data = compliance_data.get("evidences_artifacts", {})
        evidences_artifacts = EvidencesArtifacts(
            documents=evidence_data.get("documents"),
            logs=evidence_data.get("logs"),
            approvals=evidence_data.get("approvals"),
            dashboards=evidence_data.get("dashboards"),
        )
        guideline.evidences_artifacts = evidences_artifacts

        # Add and commit to the database
        db.session.add(guideline)
        db.session.commit()

        return (
            jsonify(
                {
                    "message": "Compliance data added successfully",
                    "guideline_id": guideline.id,
                }
            ),
            200,
        )

    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error: {str(e)}")
        return jsonify({"error": "Database error occurred"}), 500
    except Exception as e:
        current_app.logger.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500


@re_bp.route("/allactivity", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def get_all_activity():
    """
    Activity Page
    """
    add_to_breadcrumb(request.full_path, "All Activities")
    try:
        compliance_activities = ComplianceActivities.query.all()
        guidelines = Guidelines.query.all()
        project_names = db.session.query(Projects.project_name).distinct().all()
        print(project_names, guidelines)
        return render_template(
            "compliance_activity.html",
            compliance_activities=compliance_activities,
            guidelines=guidelines,
            project=project_names,
        )
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/get_activity", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def get_some_activity():
    """
    Activity Page
    """
    clause_id = request.args.get("clause_id")
    add_to_breadcrumb(request.full_path, "List of Activities for a Clause")
    try:
        current_app.logger.info(f"Received clause_id: {clause_id}")
        compliance_activities = ComplianceActivities.query.filter_by(
            clause_id=clause_id
        ).all()

        current_app.logger.info(
            f"Found {len(compliance_activities)} compliance activities"
        )

        # Define natural sorting function
        def natural_sort_key(activity):
            # Use activity_id for sorting (e.g., "1.1", "2.3", "10.2", etc.)
            text = activity.activity_id if activity.activity_id else ""
            if not text:
                return [float("inf")]  # Put empty/missing activity_ids at the end
            return [
                int(part) if part.isdigit() else part.lower()
                for part in re.split(r"(\d+)", str(text))
            ]

        # Sort activities using natural sorting
        sorted_activities = sorted(compliance_activities, key=natural_sort_key)

        # Debug: Print each compliance activity and its relationships
        for i, activity in enumerate(sorted_activities):
            current_app.logger.info(
                f"Activity {i}: ID={activity.id}, Clause ID={activity.clause_id}"
            )
            if hasattr(activity, "clauses"):
                current_app.logger.info(f"Activity {i} - Clause: {activity.clauses}")
                if hasattr(activity.clauses, "guideline"):
                    current_app.logger.info(
                        f"Activity {i} - Guideline: {activity.clauses.guideline}"
                    )
                    if hasattr(activity.clauses.guideline, "guideline_data"):
                        current_app.logger.info(
                            f"Activity {i} - Guideline Data: {activity.clauses.guideline.guideline_data}"
                        )
        guidelines = Guidelines.query.all()
        current_app.logger.info(f"Found {len(guidelines)} total guidelines")
        project_names = db.session.query(Projects.project_name).distinct().all()
        # Debug: Print first few guidelines structure
        for i, guideline in enumerate(guidelines[:3]):  # First 3 only
            current_app.logger.info(
                f"Guideline {i}: ID={guideline.id}, Data Type={type(guideline.guideline_data)}"
            )
            if guideline.guideline_data:
                current_app.logger.info(
                    f"Guideline {i} - Keys: {guideline.guideline_data.keys()}"
                )

        project_names = db.session.query(Projects.project_name).distinct().all()
        current_app.logger.info(f"Found {len(project_names)} distinct project names")
        return render_template(
            "compliance_activity.html",
            compliance_activities=sorted_activities,  # Use sorted_activities instead of compliance_activities
            guidelines=guidelines,
            project=project_names,
        )
    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        current_app.logger.error(f"Error type: {type(err)}")
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/activity/<int:project_id>", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def activity(project_id):
    """
    Activity Page
    """

    add_to_breadcrumb(request.full_path, "Activity Details")
    try:
        # Debug logging
        current_app.logger.info(f"Searching for project with ID: {project_id}")

        # Get project by ID instead of name
        project = Projects.query.get(project_id)

        if not project:
            current_app.logger.error(f"Project with ID {project_id} not found")
            flash(f"Project with ID {project_id} not found", "error")
            return redirect(request.referrer or url_for("audit.my_projects"))

        # Use project name from the retrieved project object
        project_name = project.project_name

        guidelines = Guidelines.query.all()
        project_names = db.session.query(Projects.project_name).distinct().all()

        # Get all compliance activities for this project
        compliance_activities = (
            db.session.query(ProjectComplianceActivity)
            .join(
                ProjectClause,
                ProjectComplianceActivity.project_clause_id == ProjectClause.id,
            )
            .join(
                ProjectGuideline,
                ProjectClause.project_guideline_id == ProjectGuideline.id,
            )
            .join(Projects, ProjectGuideline.project_id == Projects.id)
            .filter(Projects.id == project_id)
            .options(
                joinedload(ProjectComplianceActivity.project_clause).joinedload(
                    ProjectClause.project_guideline
                ),
                joinedload(ProjectComplianceActivity.project_control_activities),
            )
            .all()
        )

        # Group activities by clause and get unique clauses
        activities_by_clause = {}
        unique_clauses = {}

        for activity in compliance_activities:
            if activity.project_clause:
                clause_id = activity.project_clause.id
                if clause_id not in activities_by_clause:
                    activities_by_clause[clause_id] = []
                    unique_clauses[clause_id] = {
                        "clause": activity.project_clause,
                        "activities": [],
                    }
                activities_by_clause[clause_id].append(activity)
                unique_clauses[clause_id]["activities"].append(activity)

        # Calculate compliance status for each clause
        clause_status_info = {}
        assessment_status_info = {}
        for clause_id, activities in activities_by_clause.items():
            clause_status = get_clause_compliance_status(activities)
            clause_status_info[clause_id] = get_compliance_status_display_info(
                clause_status
            )
            assessment_status_info[clause_id] = get_assessment_status(clause_status)
        
        # ============== CALCULATE ASSESSMENT END DATE ==============
        # Check if all clauses have assessment status as "Completed"
        all_clauses_completed = True
        for clause_data in unique_clauses.values():
            clause = clause_data["clause"]
            # Get assessment status from the clause or from your logic
            clause_assessment_status = getattr(clause, 'assessment_status', 'To Be Assessed')
            if clause_assessment_status != "Completed":
                all_clauses_completed = False
                break
        
        # Determine assessment end date
        assessment_end_date = None
        if all_clauses_completed:
            # If all clauses are completed, use current date as end date
            assessment_end_date = datetime.now().date()
            current_app.logger.info(f"All clauses completed for project {project_name}. Assessment end date: {assessment_end_date}")
        else:
            # If not all completed, end date is "Present" (None)
            assessment_end_date = None
            current_app.logger.info(f"Not all clauses completed for project {project_name}. Assessment end date is Present")
        # ===========================================================
        db_assessment_end_date = project.assesment_end_date if hasattr(project, 'assesment_end_date') else None

        # Get consolidated evidence if exists
        consolidated_evidence = None
        evidence_record = ConsolidatedEvidence.query.filter_by(
            project_id=project_name
        ).first()

        if evidence_record and evidence_record.consolidate_evidence:
            if isinstance(evidence_record.consolidate_evidence, str):
                try:
                    consolidated_evidence = json.loads(
                        evidence_record.consolidate_evidence
                    )
                except json.JSONDecodeError:
                    current_app.logger.error(
                        f"Failed to parse consolidate_evidence as JSON for project {project_name}"
                    )
                    consolidated_evidence = None
            else:
                consolidated_evidence = evidence_record.consolidate_evidence

        # FIX: Process evidence files for consolidated evidence
        result_evidence = {}
        if consolidated_evidence and "grouped_evidences" in consolidated_evidence:
            result_evidence = consolidated_evidence.copy()
            for evidence_item in result_evidence["grouped_evidences"]:
                # Get evidence IDs from the required_by.evidence list
                evidence_ids = []
                if (
                    "required_by" in evidence_item
                    and "evidence" in evidence_item["required_by"]
                ):
                    evidence_ids = [
                        e.get("evidence_id")
                        for e in evidence_item["required_by"]["evidence"]
                        if e.get("evidence_id")
                    ]

                # Query evidence files for these evidence IDs
                if evidence_ids:
                    evidence_files = EvidenceFile.query.filter(
                        EvidenceFile.project_evidence_artifact_id.in_(evidence_ids)
                    ).all()

                    # Convert EvidenceFile objects to dictionaries
                    evidence_files_dict = []
                    for file in evidence_files:
                        uploaded_at = None
                        if file.uploaded_at:
                            uploaded_at = file.uploaded_at.strftime("%Y-%m-%d %H:%M")

                        evidence_files_dict.append(
                            {
                                "id": file.id,
                                "file_name": file.file_name,
                                "stored_filename": file.stored_filename,
                                "file_path": file.file_path,
                                "content_type": file.content_type,
                                "file_size": file.file_size,
                                "uploaded_at": uploaded_at,
                                "project_evidence_artifact_id": file.project_evidence_artifact_id,
                            }
                        )

                    evidence_item["evidence_files"] = evidence_files_dict
                else:
                    evidence_item["evidence_files"] = []

        # FIX: Generate evidence result with converted file objects
        result = []
        for ca in compliance_activities:
            guideline_name = None
            if ca.project_clause and ca.project_clause.project_guideline:
                gd = ca.project_clause.project_guideline.guideline_data
                if gd and "DocumentDetails" in gd:
                    guideline_name = gd["DocumentDetails"].get("DocumentName")

            clause_text = ca.project_clause.clause_text if ca.project_clause else None
            compliance_desc = ca.activity_description

            if ca.project_control_activities:
                for control in ca.project_control_activities:
                    # Load evidence files for each evidence artifact
                    evidences_with_files = []
                    if control.submitted_evidences:
                        for evidence in control.submitted_evidences:
                            # Query evidence files
                            evidence_files = EvidenceFile.query.filter_by(
                                project_evidence_artifact_id=evidence.id
                            ).all()

                            # Convert EvidenceFile objects to dictionaries
                            evidence_files_dict = []
                            for file in evidence_files:
                                uploaded_at = None
                                if file.uploaded_at:
                                    uploaded_at = file.uploaded_at.strftime(
                                        "%Y-%m-%d %H:%M"
                                    )

                                evidence_files_dict.append(
                                    {
                                        "id": file.id,
                                        "file_name": file.file_name,
                                        "stored_filename": file.stored_filename,
                                        "file_path": file.file_path,
                                        "content_type": file.content_type,
                                        "file_size": file.file_size,
                                        "uploaded_at": uploaded_at,
                                        "project_evidence_artifact_id": file.project_evidence_artifact_id,
                                    }
                                )

                            evidence_dict = {
                                "evidence_id": evidence.id,
                                "category": evidence.category,
                                "item": evidence.item,
                                "evidence_files": evidence_files_dict,
                            }
                            evidences_with_files.append(evidence_dict)

                    control_dict = {
                        "guidelines": guideline_name,
                        "clause": clause_text,
                        "compliance Activity": compliance_desc,
                        "control activity": control.activity_name,
                        "evidences": evidences_with_files,
                    }
                    result.append(control_dict)

        # Calculate project compliance status for the template
        project_compliance_status = "no-procedures"
        project_status_info = {
            "text": "To Be Assessed",
            "css_class": "bg-gray-200 text-gray-800",
        }

        if project:
            project_compliance_status = "no-procedures"
            project_status_info = get_compliance_status_display_info(
                project_compliance_status
            )

        # Create enriched data for unique clauses (not activities)
        enriched_clauses = []
        for clause_id, clause_data in unique_clauses.items():
            clause = clause_data["clause"]
            activities = clause_data["activities"]

            representative_activity = activities[0] if activities else None

            # Use the SAME calculate_clause_compliance_status function as clause_test_steps.html
            detailed_clause_status = calculate_clause_compliance_status(clause_id)

            # For backward compatibility, also calculate the old status
            old_clause_status = get_clause_compliance_status(activities)
            old_status_display = get_compliance_status_display_info(old_clause_status)
            assessment_status = get_assessment_status(old_clause_status)

            clause_obj = {
                "id": clause.id,
                "clause": clause,
                "clause_status_info": detailed_clause_status,  # This has the statistics
                "assessment_status": clause.assessment_status,
                "assessment_status_info": assessment_status,
                "representative_activity": representative_activity,
                "activities_count": len(activities),
                # Keep old structure for compatibility
                "old_clause_status": old_status_display,
            }

            enriched_clauses.append(clause_obj)
        # DEBUG: Log clause numbers before sorting
        current_app.logger.info(
            f"Clause numbers before sorting for project {project_name}:"
        )
        for i, clause_obj in enumerate(enriched_clauses):
            clause_no = clause_obj["clause"].clause_no
            current_app.logger.info(f"  {i+1}. '{clause_no}' (type: {type(clause_no)})")

        # FIX: Apply natural sorting to clauses (same as complifyre route)
        def natural_sort_key(item):
            clause_no = item["clause"].clause_no
            if clause_no is None or clause_no == "":
                return [float("inf")]
            return [
                int(part) if part.isdigit() else part.lower()
                for part in re.split(r"(\d+)", str(clause_no))
            ]

        enriched_clauses.sort(key=natural_sort_key)

        # DEBUG: Log the sorted order to verify
        current_app.logger.info(f"Sorted clauses order for project {project_name}:")
        for i, clause_obj in enumerate(enriched_clauses):
            clause_no = clause_obj["clause"].clause_no

            # Also log the sort key for debugging
            if clause_no:
                parts = re.split(r"(\d+)", str(clause_no))
                sort_key = [int(p) if p.isdigit() else p.lower() for p in parts]
                current_app.logger.info(f"  {i+1}. {clause_no} (sort_key: {sort_key})")
            else:
                current_app.logger.info(f"  {i+1}. EMPTY/None (sort_key: [inf])")

        # Enhanced debug logging for consolidated evidence
        current_app.logger.info(
            f"Activity route - Project {project_name} - Consolidated evidence type: {type(consolidated_evidence)}"
        )
        if consolidated_evidence and isinstance(consolidated_evidence, dict):
            grouped_evidences = consolidated_evidence.get("grouped_evidences", [])
            current_app.logger.info(
                f"Activity route - Project {project_name} - Grouped evidences length: {len(grouped_evidences)}"
            )

        # Calculate statistics for dashboard
        def calculate_clause_statistics(compliance_clauses):
            """Calculate statistics for the dashboard pie charts."""
            total_clauses = len(compliance_clauses)

            # Applicability Statistics
            applicable_clauses = sum(
                1 for clause in compliance_clauses if clause["clause"].applicability
            )
            not_applicable_clauses = total_clauses - applicable_clauses

            # Assessment Status Statistics (only for applicable clauses)
            completed_assessments = sum(
                1
                for clause in compliance_clauses
                if clause["clause"].applicability
                and clause.get("assessment_status") == "Completed"
            )
            to_be_assessed = applicable_clauses - completed_assessments

            # Compliance Status Statistics (only for completed assessments)
            compliant_clauses = 0
            partially_compliant_clauses = 0
            non_compliant_clauses = 0

            for clause in compliance_clauses:
                if (
                    clause["clause"].applicability
                    and clause.get("assessment_status") == "Completed"
                ):
                    status = clause["clause_status_info"]["text"]
                    if status == "Compliant":
                        compliant_clauses += 1
                    elif status == "Partially Compliant":
                        partially_compliant_clauses += 1
                    elif status == "Non-Compliant":
                        non_compliant_clauses += 1

            return {
                "total_clauses": total_clauses,
                "applicability": {
                    "applicable": applicable_clauses,
                    "not_applicable": not_applicable_clauses,
                    "percentage_applicable": (
                        round((applicable_clauses / total_clauses * 100), 1)
                        if total_clauses > 0
                        else 0
                    ),
                    "percentage_not_applicable": (
                        round((not_applicable_clauses / total_clauses * 100), 1)
                        if total_clauses > 0
                        else 0
                    ),
                },
                "assessment": {
                    "completed": completed_assessments,
                    "to_be_assessed": to_be_assessed,
                    "percentage_completed": (
                        round((completed_assessments / applicable_clauses * 100), 1)
                        if applicable_clauses > 0
                        else 0
                    ),
                    "percentage_to_be_assessed": (
                        round((to_be_assessed / applicable_clauses * 100), 1)
                        if applicable_clauses > 0
                        else 0
                    ),
                },
                "compliance": {
                    "compliant": compliant_clauses,
                    "partially_compliant": partially_compliant_clauses,
                    "non_compliant": non_compliant_clauses,
                    "total_assessed": completed_assessments,
                    "percentage_compliant": (
                        round((compliant_clauses / completed_assessments * 100), 1)
                        if completed_assessments > 0
                        else 0
                    ),
                    "percentage_partially_compliant": (
                        round(
                            (partially_compliant_clauses / completed_assessments * 100),
                            1,
                        )
                        if completed_assessments > 0
                        else 0
                    ),
                    "percentage_non_compliant": (
                        round((non_compliant_clauses / completed_assessments * 100), 1)
                        if completed_assessments > 0
                        else 0
                    ),
                },
            }

        # Calculate statistics
        clause_statistics = calculate_clause_statistics(enriched_clauses)

        # ============== CALCULATE EVIDENCE RECEIVED STATISTICS ==============
        # For each clause, determine if evidence is received (ALL applicable activities have admissible evidence)
        clauses_with_evidence = 0
        clauses_without_evidence = 0
        severity_counts = {
            'Critical': 0,
            'Major': 0,
            'Significant': 0,
            'Minor': 0,
            'No findings noted': 0
        }
        
        for clause_data in enriched_clauses:
            clause = clause_data["clause"]
            
            # Skip if clause is not applicable
            if not clause.applicability:
                continue
            
            # Get all control activities for this clause through the proper join path
            control_activities = (
                db.session.query(ProjectControlActivity)
                .join(
                    ProjectComplianceActivity,
                    ProjectControlActivity.project_compliance_activity_id == ProjectComplianceActivity.id
                )
                .filter(ProjectComplianceActivity.project_clause_id == clause.id)
                .filter(ProjectComplianceActivity.applicability == True)  # Only applicable activities
                .all()
            )
            
            # If there are no applicable activities, skip
            if not control_activities:
                continue
            
            # Check if ALL applicable activities have admissible evidence
            all_activities_have_evidence = True
            activities_with_evidence_count = 0
            total_applicable_activities = len(control_activities)
            
            for activity in control_activities:
                # Check if activity has evidence received (admissible and strong)
                evidence_received = (
                    activity.evidence_admissibility_decision == "Yes" and 
                    activity.evidence_quality_rating == "STRONG"
                )
                
                if evidence_received:
                    activities_with_evidence_count += 1
                else:
                    # If any activity lacks evidence, the clause fails the ALL condition
                    all_activities_have_evidence = False
                    # Log which activity is missing evidence for debugging
                    current_app.logger.info(f"Clause {clause.clause_no} - Activity {activity.activity_code} missing evidence: Admissibility={activity.evidence_admissibility_decision}, Quality={activity.evidence_quality_rating}")
            
            # Determine clause evidence status based on ALL activities having evidence
            if all_activities_have_evidence:
                clauses_with_evidence += 1
                evidence_status = "YES"
            else:
                clauses_without_evidence += 1
                evidence_status = "NO"
            
            # Log for debugging
            current_app.logger.info(f"Clause {clause.clause_no}: {activities_with_evidence_count}/{total_applicable_activities} activities with evidence - ALL have evidence: {all_activities_have_evidence} -> Evidence Status: {evidence_status}")
            
            # ============== CALCULATE SEVERITY FOR THIS CLAUSE ==============
            # Get the overall severity for this clause (highest severity across all applicable activities)
            highest_severity = 'No findings noted'
            highest_score = 0
            
            severity_hierarchy = {
                'Critical': 5,
                'Major': 4,
                'Significant': 3,
                'Minor': 2,
                'No findings noted': 1
            }
            
            for activity in control_activities:
                activity_severity = activity.overall_severity_classification
                
                # Default to 'No findings noted' if nothing found and activity is compliant
                if not activity_severity or activity_severity == 'Not Classified':
                    if activity.compliant_status == 'Compliant':
                        activity_severity = 'No findings noted'
                    else:
                        activity_severity = 'Not Classified'
                
                if activity_severity and activity_severity != 'Not Classified':
                    severity_score = severity_hierarchy.get(activity_severity, 0)
                    if severity_score > highest_score:
                        highest_score = severity_score
                        highest_severity = activity_severity
            
            # Count this clause's severity (only count if it's in our dictionary)
            if highest_severity in severity_counts:
                severity_counts[highest_severity] += 1
            else:
                # Default to 'No findings noted' for any unclassified severity
                severity_counts['No findings noted'] += 1
        
        # Calculate percentages
        total_applicable_clauses = clauses_with_evidence + clauses_without_evidence
        evidence_percentage = 0
        if total_applicable_clauses > 0:
            evidence_percentage = round((clauses_with_evidence / total_applicable_clauses) * 100)
        
        evidence_stats = {
            'with_evidence': clauses_with_evidence,
            'without_evidence': clauses_without_evidence,
            'total': total_applicable_clauses,
            'percentage': evidence_percentage,
            'color_class': 'text-green-600' if evidence_percentage >= 75 else 
                           'text-yellow-600' if evidence_percentage >= 50 else 
                           'text-orange-600' if evidence_percentage >= 25 else 'text-red-600',
            'progress_class': 'bg-green-500' if evidence_percentage >= 75 else 
                              'bg-yellow-500' if evidence_percentage >= 50 else 
                              'bg-orange-500' if evidence_percentage >= 25 else 'bg-red-500'
        }

        # Calculate assessment status statistics for the status bar
        assessment_status_stats = {
            'Completed': clause_statistics['assessment']['completed'],
            'To Be Assessed': clause_statistics['assessment']['to_be_assessed'],
            'total': clause_statistics['applicability']['applicable']
        }
        
        # Calculate compliance status statistics
        compliance_status_stats = {
            'Compliant': clause_statistics['compliance']['compliant'],
            'Partially Compliant': clause_statistics['compliance']['partially_compliant'],
            'Non-Compliant': clause_statistics['compliance']['non_compliant'],
            'total': clause_statistics['compliance']['total_assessed'],
            'percentage_compliant': clause_statistics['compliance']['percentage_compliant']
        }
        
        # Calculate evaluated activities count
        evaluated_activities_count = clause_statistics['assessment']['completed']
        total_applicable_activities = clause_statistics['applicability']['applicable']


        from datetime import datetime
        current_time = datetime.now()
        
        # ============== CALCULATE SEVERITY STATISTICS ==============
        severity_stats = {
            'Critical': severity_counts.get('Critical', 0),
            'Major': severity_counts.get('Major', 0),
            'Significant': severity_counts.get('Significant', 0),
            'Minor': severity_counts.get('Minor', 0),
            'No findings noted': severity_counts.get('No findings noted', 0)
        }
        
        # Determine overall project severity (highest severity found)
        overall_project_severity = 'No findings noted'
        if severity_stats['Critical'] > 0:
            overall_project_severity = 'Critical'
        elif severity_stats['Major'] > 0:
            overall_project_severity = 'Major'
        elif severity_stats['Significant'] > 0:
            overall_project_severity = 'Significant'
        elif severity_stats['Minor'] > 0:
            overall_project_severity = 'Minor'
        
        severity_color_class = {
            'Critical': 'bg-red-600 text-white',
            'Major': 'bg-orange-500 text-white',
            'Significant': 'bg-yellow-500 text-gray-900',
            'Minor': 'bg-blue-400 text-white',
            'No findings noted': 'bg-green-500 text-white'
        }.get(overall_project_severity, 'bg-gray-400 text-white')
        # Render my_projects_new.html instead of project_activity.html
        return render_template(
            "my_projects_new.html",
            compliance_clauses=enriched_clauses,
            guidelines=guidelines,
            project_name=project_name,
            project=project,
            consolidated_evidence=result_evidence,
            evidence_result=result,
            project_compliance_status=project_compliance_status,
            project_status_info=project_status_info,
            clause_statistics=clause_statistics,
            now=datetime.now(),

            assessment_start_date=project.assesment_start_date if project else None,
            assessment_end_date=assessment_end_date,
            db_assessment_end_date=db_assessment_end_date,
            all_clauses_completed=all_clauses_completed,
            evidence_stats=evidence_stats,
            severity_stats=severity_stats,
            overall_project_severity=overall_project_severity,
            severity_color_class=severity_color_class,
            assessment_status_stats=assessment_status_stats,
            compliance_status_stats=compliance_status_stats,
            evaluated_activities_count=evaluated_activities_count,
            total_applicable_activities=total_applicable_activities,
            current_time=current_time,
        )

    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}", exc_info=True)
        return jsonify({"error": f"Internal server error {err}"}), 500



@re_bp.route("/get_clause_statistics/<int:project_id>", methods=["GET"])
def get_clause_statistics(project_id):
    """
    Get updated clause statistics for charts
    """
    try:
        # Get project
        project = Projects.query.get(project_id)
        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Get all clauses for this project
        clauses = (
            ProjectClause.query.join(
                ProjectGuideline,
                ProjectClause.project_guideline_id == ProjectGuideline.id,
            )
            .filter(ProjectGuideline.project_id == project_id)
            .all()
        )

        # Calculate statistics
        total_clauses = len(clauses)

        # Applicability Statistics
        applicable_clauses = sum(1 for clause in clauses if clause.applicability)
        not_applicable_clauses = total_clauses - applicable_clauses

        # Assessment Status Statistics (only for applicable clauses)
        # For assessment, we need to check if activities exist and have been evaluated
        completed_assessments = 0
        to_be_assessed = 0

        for clause in clauses:
            if clause.applicability:
                # Check if this clause has been assessed
                # Get compliance status for the clause
                clause_status_data = calculate_clause_compliance_status(clause.id)
                status_text = clause_status_data.get("text", "To Be Assessed")

                if status_text != "To Be Assessed":
                    completed_assessments += 1
                else:
                    to_be_assessed += 1

        # Compliance Status Statistics (only for completed assessments)
        compliant_clauses = 0
        partially_compliant_clauses = 0
        non_compliant_clauses = 0

        for clause in clauses:
            if clause.applicability:
                clause_status_data = calculate_clause_compliance_status(clause.id)
                status_text = clause_status_data.get("text", "To Be Assessed")

                if status_text == "Compliant":
                    compliant_clauses += 1
                elif status_text == "Partially Compliant":
                    partially_compliant_clauses += 1
                elif status_text == "Non-Compliant":
                    non_compliant_clauses += 1

        statistics = {
            "total_clauses": total_clauses,
            "applicability": {
                "applicable": applicable_clauses,
                "not_applicable": not_applicable_clauses,
                "percentage_applicable": (
                    round((applicable_clauses / total_clauses * 100), 1)
                    if total_clauses > 0
                    else 0
                ),
                "percentage_not_applicable": (
                    round((not_applicable_clauses / total_clauses * 100), 1)
                    if total_clauses > 0
                    else 0
                ),
            },
            "assessment": {
                "completed": completed_assessments,
                "to_be_assessed": to_be_assessed,
                "percentage_completed": (
                    round((completed_assessments / applicable_clauses * 100), 1)
                    if applicable_clauses > 0
                    else 0
                ),
                "percentage_to_be_assessed": (
                    round((to_be_assessed / applicable_clauses * 100), 1)
                    if applicable_clauses > 0
                    else 0
                ),
            },
            "compliance": {
                "compliant": compliant_clauses,
                "partially_compliant": partially_compliant_clauses,
                "non_compliant": non_compliant_clauses,
                "total_assessed": completed_assessments,
                "percentage_compliant": (
                    round((compliant_clauses / completed_assessments * 100), 1)
                    if completed_assessments > 0
                    else 0
                ),
                "percentage_partially_compliant": (
                    round(
                        (partially_compliant_clauses / completed_assessments * 100), 1
                    )
                    if completed_assessments > 0
                    else 0
                ),
                "percentage_non_compliant": (
                    round((non_compliant_clauses / completed_assessments * 100), 1)
                    if completed_assessments > 0
                    else 0
                ),
            },
        }

        return jsonify({"status": "success", "statistics": statistics})

    except Exception as err:
        current_app.logger.error(
            f"Error getting clause statistics: {str(err)}", exc_info=True
        )
        return jsonify({"error": f"Internal server error {err}"}), 500



@re_bp.route("/activity_perform", methods=["POST"])
## @role_required()
def activity_perform():
    """
    Activity Page
    """
    try:
        pdf_service = PDFService()
        data = request.get_json()
        clauses = Clauses.query.filter_by(id=data.get("clause_id")).first()
        print(clauses, data.get("clause_text"))
        guideline_id = clauses.guideline_id
        if not guideline_id:
            raise ValueError("Guideline ID is required")
        # Query the database for the guideline and its associated clauses
        guideline = Guidelines.query.filter_by(id=guideline_id).first()
        if not guideline:
            raise ValueError("Guideline not found")
        file_url = File.query.filter_by(id=guideline.file_id).first()
        print(file_url)
        # Prepare data for rendering
        url = file_url.path
        text = pdf_service.extract_text_from_pdf(url)
        json_data = pdf_service.retrive_activity(clauses.clause_text, data, text)
        compliance_data = json.loads(f"""{json_data}""")
        activity_data = HowToPerformActivity(
            activity_id=data.get("id"), data=compliance_data
        )
        db.session.add(activity_data)
        db.session.commit()
        return (
            jsonify(
                {
                    "message": "Activity data added successfully",
                    "guideline_id": compliance_data,
                }
            ),
            200,
        )
    except SQLAlchemyError as e:
        db.session.rollback()
        current_app.logger.error(f"Database error: {str(e)}")


@re_bp.route("/generate_all_activities", methods=["POST"])
def generate_all_activities():
    try:
        pdf_service = PDFService()
        data = request.get_json()
        all_activities = data.get("activities", [])
        print(all_activities)
        for activity in all_activities:
            print(activity)
            clauses = Clauses.query.filter_by(id=activity.get("clause_id")).first()
            print("clause", clauses)
            if not clauses:
                continue
            guideline_id = clauses.guideline_id
            print("guideline_id", guideline_id)
            if not guideline_id:
                continue
            guideline = Guidelines.query.filter_by(id=guideline_id).first()
            print("guide", guideline)
            if not guideline:
                continue
            file_url = File.query.filter_by(id=guideline.file_id).first()
            print("file", file_url)
            if not file_url:
                continue

            url = file_url.path
            text = pdf_service.extract_text_from_pdf(url)
            json_data = pdf_service.retrive_activity(
                clauses.clause_text, activity, text
            )
            compliance_data = json.loads(f"""{json_data}""")

            activity_data = HowToPerformActivity(
                activity_id=activity.get("id"), data=compliance_data
            )
            db.session.add(activity_data)

        db.session.commit()
        return (
            jsonify({"message": "All activity instructions generated successfully."}),
            200,
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(
            f"Error generating all activity instructions: {str(e)}"
        )
        return jsonify({"message": "Error occurred during batch generation."}), 500


@re_bp.route("/how_to_perform")
def how_to_perform():
    try:
        # Fetch all how-to-perform activities from the database
        how_to_perform_activities = HowToPerformActivity.query.all()
        for item in how_to_perform_activities:
            print(item.data)
        return render_template(
            "activity_mapping.html", how_to_perform_activities=how_to_perform_activities
        )
    except Exception as e:
        current_app.logger.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500


@re_bp.route("/evidence_artifacts/<int:activity_id>")
def eviences_artifacts(activity_id):
    try:
        # Fetch all how-to-perform activities from the database
        how_to_perform_activities = HowToPerformActivity.query.filter_by(
            activity_id=activity_id
        ).all()
        for item in how_to_perform_activities:
            print(item.data)
        return render_template(
            "activity_mapping.html", how_to_perform_activities=how_to_perform_activities
        )
    except Exception as e:
        current_app.logger.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "An unexpected error occurred"}), 500


# @re_bp.route(' /<int:activity_id>')
# def test_evidence_artifacts(activity_id):
#     try:
#         # Fetch all how-to-perform activities from the database
#         how_to_perform_activities = TestProcedures.query.filter_by(activity_id=activity_id).all()
#         for item in how_to_perform_activities:
#             print(item.data)
#         return render_template('test_artifacts.html', how_to_perform_activities=how_to_perform_activities)
#     except Exception as e:
#         current_app.logger.error(f"Unexpected error: {str(e)}")
#         return jsonify({"error": "An unexpected error occurred"}), 500


@re_bp.route("/test_evidence_artifacts/<int:activity_id>")
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def test_evidence_artifacts(activity_id):
    """
    Displays the detailed testing artifacts for a single project control activity.
    """
    add_to_breadcrumb(request.full_path, "Test Procedure")
    print("Activity Ids", activity_id)
    try:
        # The query now starts from the project-specific control activity table.
        project_control = (
            db.session.query(ProjectControlActivity)
            .filter_by(id=activity_id)
            .options(
                joinedload(ProjectControlActivity.project_test_procedure).joinedload(
                    ProjectTestSteps.project_documents
                ),
                joinedload(ProjectControlActivity.project_test_procedure)
                .joinedload(ProjectTestSteps.project_interview)
                .joinedload(ProjectInterview.project_roles),
                joinedload(ProjectControlActivity.project_test_procedure)
                .joinedload(ProjectTestSteps.project_interview)
                .joinedload(ProjectInterview.project_questions),
                joinedload(ProjectControlActivity.submitted_evidences),
                joinedload(ProjectControlActivity.project_test_procedure).joinedload(
                    ProjectTestSteps.test_procedure_files
                ),
                # FIX: Eagerly load the compliance activity and its clause relationship
                joinedload(
                    ProjectControlActivity.project_compliance_activity
                ).joinedload(ProjectComplianceActivity.project_clause),
                # ADD: Load project relationship to get project data
                joinedload(ProjectControlActivity.project_compliance_activity)
                .joinedload(ProjectComplianceActivity.project_clause)
                .joinedload(ProjectClause.project_guideline)
                .joinedload(ProjectGuideline.project),
            )
            .first_or_404()
        )

        # DEBUG: Print what we're getting
        print(f"Control Activity Name: '{project_control.activity_name}'")
        print(f"Control Activity Description: '{project_control.activity_description}'")
        print(
            f"Parent Compliance Activity Description: '{project_control.project_compliance_activity.activity_description if project_control.project_compliance_activity else 'None'}'"
        )

        # Use the parent compliance activity's description as the activity name
        display_activity_name = (
            project_control.project_compliance_activity.activity_description
        )

        print(f"Final Display Activity Name: '{display_activity_name}'")
        print(project_control)

        # FIX: Get clause_id safely through relationships
        clause_id = None
        project = None
        if (
            project_control.project_compliance_activity
            and project_control.project_compliance_activity.project_clause
        ):
            clause_id = project_control.project_compliance_activity.project_clause.id

            # Get project from the relationship chain
            if (
                project_control.project_compliance_activity.project_clause.project_guideline
            ):
                project = (
                    project_control.project_compliance_activity.project_clause.project_guideline.project
                )

        print(f"Clause ID: {clause_id}")
        print(f"Project: {project}")

        p_test_procedure = project_control.project_test_procedure
        p_interview = p_test_procedure.project_interview if p_test_procedure else None

        document_reviews = (
            [doc.document_name for doc in p_test_procedure.project_documents]
            if p_test_procedure
            else []
        )

        interview_data = {
            "roles": (
                [role.role for role in p_interview.project_roles] if p_interview else []
            ),
            "key_questions": (
                [
                    {
                        "id": q.id,
                        "question": q.question,
                        "answer": q.answer,
                    }
                    for q in p_interview.project_questions
                ]
                if p_interview
                else []
            ),
        }

        evidence_data = {}
        for p_evidence in project_control.submitted_evidences:
            evidence_entry = {
                "id": p_evidence.id,
                "item": p_evidence.item,
                "evidence_ans": p_evidence.evidence_text,
                "files": p_evidence.evidence_file_path,
            }
            evidence_data.setdefault(p_evidence.category, []).append(evidence_entry)

        # Process file attachments for walkthrough and sampling
        walkthrough_files = []
        sampling_files = []

        if p_test_procedure and p_test_procedure.test_procedure_files:
            for file in p_test_procedure.test_procedure_files:
                if file.field_type == "walkthrough_files":
                    walkthrough_files.append(
                        {
                            "id": file.id,
                            "filename": file.filename,
                            "file_size": file.file_size,
                            "upload_date": file.upload_date,
                        }
                    )
                elif file.field_type == "sampling_files":
                    sampling_files.append(
                        {
                            "id": file.id,
                            "filename": file.filename,
                            "file_size": file.file_size,
                            "upload_date": file.upload_date,
                        }
                    )

        # FIX: Only query consolidated summary if we have a valid clause_id
        consolidated_test_data = None
        if clause_id:
            consolidated_summary = (
                ConsolidatedTestSummary.query.filter_by(clause_id=clause_id)
                .order_by(ConsolidatedTestSummary.generated_at.desc())
                .first()
            )

            if consolidated_summary:
                try:
                    consolidated_test_data = json.loads(
                        consolidated_summary.consolidated_summary
                    )
                except json.JSONDecodeError:
                    consolidated_test_data = {
                        "consolidated_summary": consolidated_summary.consolidated_summary
                    }

        # All data is now sourced from the 'project_control' instance object
        activity_data = {
            "activity_id": project_control.id,
            "activity_code": project_control.activity_code,
            "activity_name": display_activity_name,
            "activity_description": project_control.activity_description,
            "objective": project_control.objective,
            "owner": project_control.owner,
            "control_type": project_control.control_type,
            "frequency": project_control.frequency,
            "sampling_guidance": project_control.sampling_guidance,
            "auditor_observation": project_control.auditor_observation,
            "findings": project_control.findings,
            "impact": project_control.impact,
            "severity": project_control.severity,
            "recommendations": project_control.recommendations,
            "observation": project_control.auditor_observation,
            "reviewer_notes": project_control.reviewer_notes,
            "explain_test_procedure": project_control.explain_test_procedure,

            # **NEW FIELDS - ADD THESE**
            "evidence_admissibility_decision": project_control.evidence_admissibility_decision,
            "evidence_quality_rating": project_control.evidence_quality_rating,
            "reason_for_inadmissibility": project_control.reason_for_inadmissibility,
            "required_effectiveness_design": project_control.required_effectiveness_design,
            "required_effectiveness_implementation": project_control.required_effectiveness_implementation,
            "required_effectiveness_operating": project_control.required_effectiveness_operating,
            "detailed_control_testing_results": project_control.detailed_control_testing_results,
            "severity_classification_per_finding": project_control.severity_classification_per_finding,
            "overall_severity_classification": project_control.overall_severity_classification,
            "test_procedure": {
                "Walkthrough": (
                    p_test_procedure.walkthrough if p_test_procedure else None
                ),
                "sampling": p_test_procedure.sampling if p_test_procedure else None,
                "additional_walkthrough": (
                    p_test_procedure.additional_walkthrough
                    if p_test_procedure
                    else None
                ),
                "additional_sampling": (
                    p_test_procedure.additional_sampling if p_test_procedure else None
                ),
                "walkthrough_files": walkthrough_files,
                "sampling_files": sampling_files,
                "review_of_documentation": document_reviews,
                "interviews": interview_data,
            },
            "evidences_artifacts_needed": evidence_data,
            "compliant_status": project_control.compliant_status,
            "control_finding": project_control.control_findings,
            "control_recomend": project_control.control_recommendation,
            "consolidated_test_summary": consolidated_test_data,
        }

        return render_template(
            "test_artifacts.html",
            how_to_perform_activities=[{"data": activity_data}],
            consolidated_test_summary=consolidated_test_data,
            project=project,
            clause_id=clause_id,
        )

    except Exception as e:
        current_app.logger.error(
            f"Unexpected error in test_evidence_artifacts: {str(e)}"
        )
        return redirect(request.referrer)



@re_bp.route("/delete_test_procedure_main_content", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def delete_test_procedure_main_content():
    """
    Delete main walkthrough or sampling content for a test procedure.
    """
    try:
        data = request.get_json()
        activity_id = data.get('activity_id')
        field = data.get('field')  # 'walkthrough' or 'sampling'
        
        if not activity_id or not field:
            return jsonify({
                'success': False,
                'message': 'Missing required fields'
            }), 400
            
        # Validate field name for security
        if field not in ['walkthrough', 'sampling']:
            return jsonify({
                'success': False,
                'message': 'Invalid field specified'
            }), 400
            
        # Get the project control activity
        project_control = ProjectControlActivity.query.get_or_404(activity_id)
        
        # Get the associated test procedure
        test_procedure = project_control.project_test_procedure
        
        if not test_procedure:
            return jsonify({
                'success': False,
                'message': 'No test procedure found for this activity'
            }), 404
            
        # Set the field to empty string or None
        setattr(test_procedure, field, '')
        
        # Commit the changes
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{field.title()} content deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting test procedure main content: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error deleting content: {str(e)}'
        }), 500

@re_bp.route("/delete_test_procedure_content", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def delete_test_procedure_content():
    try:
        data = request.get_json()
        activity_id = data.get('activity_id')
        field = data.get('field')  # 'additional_walkthrough' or 'additional_sampling'
        
        if not activity_id or not field:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
            
        if field not in ['additional_walkthrough', 'additional_sampling']:
            return jsonify({'success': False, 'message': 'Invalid field specified'}), 400
            
        project_control = ProjectControlActivity.query.get_or_404(activity_id)
        test_procedure = project_control.project_test_procedure
        
        if not test_procedure:
            return jsonify({'success': False, 'message': 'No test procedure found'}), 404
            
        setattr(test_procedure, field, None)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'{field} content deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting additional content: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500



@re_bp.route("/evaluate_clause_activities", methods=["POST"])
@login_required
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def evaluate_clause_activities():
    """
    Evaluate all applicable activities for a clause using consolidated evidence
    and AI-powered detailed analysis
    """
    try:
        clause_id = request.form.get("clause_id", type=int)
        project_id = request.form.get("project_id", type=int)

        if not clause_id:
            return jsonify({"success": False, "error": "Clause ID required"}), 400

        # Get the clause and project
        clause = ProjectClause.query.get_or_404(clause_id)
        project = Projects.query.get_or_404(project_id)

        # Get all applicable control activities for this clause
        applicable_activities = []

        # Get compliance activities for this clause
        compliance_activities = ProjectComplianceActivity.query.filter_by(
            project_clause_id=clause_id, applicability=True  # Only applicable ones
        ).all()

        for compliance_activity in compliance_activities:
            # Get control activities for each compliance activity
            control_activities = ProjectControlActivity.query.filter_by(
                project_compliance_activity_id=compliance_activity.id
            ).all()

            for control_activity in control_activities:
                applicable_activities.append(
                    {
                        "id": control_activity.id,
                        "activity_code": control_activity.activity_code,
                        "activity_name": control_activity.activity_description
                        or control_activity.activity_name,
                        "compliance_activity_id": compliance_activity.id,
                        "activity_object": control_activity,  # Keep reference to the object
                    }
                )

        if not applicable_activities:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "No applicable activities found for this clause",
                    }
                ),
                404,
            )

        # Get consolidated evidence for the project
        consolidated_evidence = None
        evidence_record = ConsolidatedEvidence.query.filter_by(
            project_id=project.project_name
        ).first()

        if evidence_record and evidence_record.consolidate_evidence:
            try:
                if isinstance(evidence_record.consolidate_evidence, str):
                    consolidated_evidence = json.loads(
                        evidence_record.consolidate_evidence
                    )
                else:
                    consolidated_evidence = evidence_record.consolidate_evidence
            except json.JSONDecodeError:
                current_app.logger.error(
                    f"Failed to parse consolidate_evidence for project {project.project_name}"
                )
                consolidated_evidence = None

        # Evaluate each applicable activity using AI
        evaluated_count = 0
        evaluation_results = []

        for activity in applicable_activities:
            try:
                # Collect all evidence for this activity
                evidence_data = []

                # 1. Get submitted evidences from the activity itself
                submitted_evidences = []
                control_activity_obj = activity["activity_object"]
                for evidence in control_activity_obj.submitted_evidences:
                    submitted_evidences.append(
                        {
                            "item": evidence.item,
                            "evidence_text": evidence.evidence_text,
                            "files": evidence.evidence_file_path,
                            "evidence_type": "submitted",
                        }
                    )

                # 2. Get evidence from consolidated evidence
                consolidated_files = []
                if (
                    consolidated_evidence
                    and "grouped_evidences" in consolidated_evidence
                ):
                    for evidence_group in consolidated_evidence["grouped_evidences"]:
                        if (
                            evidence_group.get("required_by")
                            and "activity_ids" in evidence_group["required_by"]
                        ):
                            if (
                                activity["id"]
                                in evidence_group["required_by"]["activity_ids"]
                            ):
                                # Get evidence files for this group
                                evidence_ids = [
                                    e.get("evidence_id")
                                    for e in evidence_group["required_by"].get(
                                        "evidence", []
                                    )
                                    if e.get("evidence_id")
                                ]

                                if evidence_ids:
                                    files = EvidenceFile.query.filter(
                                        EvidenceFile.project_evidence_artifact_id.in_(
                                            evidence_ids
                                        )
                                    ).all()

                                    for file in files:
                                        consolidated_files.append(
                                            {
                                                "filename": file.file_name,
                                                "file_path": file.file_path,
                                                "content_type": file.content_type,
                                                "size": file.file_size,
                                                "evidence_type": "consolidated",
                                            }
                                        )

                # 3. Get test procedure information if available
                test_procedure_info = None
                if control_activity_obj.project_test_procedure:
                    test_proc = control_activity_obj.project_test_procedure
                    test_procedure_info = {
                        "additional_walkthrough": getattr(
                            test_proc, "additional_walkthrough", None
                        ),
                        "additional_sampling": getattr(
                            test_proc, "additional_sampling", None
                        ),
                        "has_files": bool(
                            test_proc.test_procedure_files
                            and len(test_proc.test_procedure_files) > 0
                        ),
                    }

                # 4. Call the enhanced evaluation function
                evaluation_result = evaluate_single_activity_ai(
                    control_activity=control_activity_obj,
                    submitted_evidences=submitted_evidences,
                    consolidated_files=consolidated_files,
                    test_procedure_info=test_procedure_info,
                    activity_context={
                        "clause_no": clause.clause_no,
                        "clause_text": clause.clause_text,
                        "project_name": project.project_name,
                        "project_id": project_id,
                    },
                )

                if evaluation_result["success"]:
                    evaluated_count += 1
                    evaluation_results.append(
                        {
                            "activity_id": activity["id"],
                            "activity_code": activity["activity_code"],
                            "activity_name": activity["activity_name"],
                            "result": evaluation_result,
                        }
                    )

                    # Update the control activity with AI evaluation results
                    if evaluation_result.get("compliant_status"):
                        control_activity_obj.compliant_status = evaluation_result[
                            "compliant_status"
                        ]
                        control_activity_obj.auditor_observation = (
                            evaluation_result.get("observation", "")
                        )
                        control_activity_obj.findings = evaluation_result.get(
                            "findings", ""
                        )
                        control_activity_obj.recommendations = evaluation_result.get(
                            "recommendations", ""
                        )
                        control_activity_obj.updated_at = datetime.utcnow()

                        # Log the evaluation
                        current_app.logger.info(
                            f"AI evaluated activity {activity['activity_code']} - "
                            f"Status: {evaluation_result['compliant_status']}"
                        )

            except Exception as e:
                current_app.logger.error(
                    f"Error evaluating activity {activity['id']}: {str(e)}"
                )
                # Fallback to basic evaluation if AI fails
                try:
                    basic_result = evaluate_single_activity_ai(
                        activity_id=activity["id"],
                        evidence_files=consolidated_files,
                        activity_context={
                            "clause_no": clause.clause_no,
                            "clause_text": clause.clause_text,
                            "project_name": project.project_name,
                        },
                    )
                    if basic_result["success"]:
                        evaluated_count += 1
                        control_activity_obj.compliant_status = basic_result[
                            "compliant_status"
                        ]
                        control_activity_obj.auditor_observation = basic_result.get(
                            "observation", ""
                        )
                        control_activity_obj.findings = basic_result.get("findings", "")
                        control_activity_obj.recommendations = basic_result.get(
                            "recommendations", ""
                        )
                        control_activity_obj.updated_at = datetime.utcnow()
                except Exception as fallback_error:
                    current_app.logger.error(
                        f"Fallback evaluation also failed: {str(fallback_error)}"
                    )
                continue

        # Commit all updates
        db.session.commit()

        # Recalculate clause compliance status
        clause_status_info = calculate_clause_compliance_status(clause_id)

        # Update clause assessment status if all activities evaluated
        if evaluated_count == len(applicable_activities):
            clause.assessment_status = "Completed"
            clause.overall_compliance_status = clause_status_info["text"]
            db.session.commit()

        # Prepare response data
        response_data = {
            "success": True,
            "message": f"Successfully evaluated {evaluated_count} out of {len(applicable_activities)} applicable activities using AI analysis",
            "evaluated_count": evaluated_count,
            "total_applicable": len(applicable_activities),
            "clause_status": {
                "text": clause_status_info["text"],
                "css_class": clause_status_info["css_class"],
            },
            "statistics": clause_status_info.get("statistics", {}),
            "redirect_url": None,
            "details": {
                "ai_evaluated": evaluated_count,
                "activities": [
                    {
                        "id": activity["id"],
                        "code": activity["activity_code"],
                        "name": activity["activity_name"],
                        "status": (
                            activity["activity_object"].compliant_status
                            if activity["activity_object"].compliant_status
                            else "not-evaluated"
                        ),
                    }
                    for activity in applicable_activities
                ],
            },
        }

        # If at least one activity was evaluated, provide redirect to first one
        if evaluation_results and evaluation_results[0]["activity_id"]:
            response_data["redirect_url"] = url_for(
                "re.test_evidence_artifacts",
                activity_id=evaluation_results[0]["activity_id"],
            )

        return jsonify(response_data)

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error in evaluate_clause_activities: {str(e)}")
        return (
            jsonify({"success": False, "error": f"Internal server error: {str(e)}"}),
            500,
        )


@re_bp.route("/update_test_procedure", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def update_test_procedure():
    """
    Updates the walkthrough, sampling, or additional information fields of a test procedure.
    """
    try:
        data = request.get_json()
        activity_id = data.get("activity_id")
        field = data.get("field")
        value = data.get("value")

        print(f"DEBUG: Received - activity_id: {activity_id}, field: {field}")  # Debug

        if not all([activity_id, field, value is not None]):
            return jsonify({"success": False, "message": "Missing required parameters"})

        # Get the project control activity
        project_control = ProjectControlActivity.query.filter_by(id=activity_id).first()

        if not project_control:
            return jsonify({"success": False, "message": "Activity not found"})

        # Get or create the test procedure
        if not project_control.project_test_procedure:
            # Create a new test procedure if it doesn't exist
            new_test_procedure = ProjectTestSteps()
            db.session.add(new_test_procedure)
            project_control.project_test_procedure = new_test_procedure
            db.session.flush()  # This ensures we get an ID

        # Update the appropriate field
        if field == "walkthrough":
            project_control.project_test_procedure.walkthrough = value
        elif field == "sampling":
            project_control.project_test_procedure.sampling = value
        elif field == "additional_walkthrough":
            project_control.project_test_procedure.additional_walkthrough = value
        elif field == "additional_sampling":
            project_control.project_test_procedure.additional_sampling = value
        else:
            return jsonify({"success": False, "message": "Invalid field specified"})

        db.session.commit()

        # Return the saved value to confirm it was saved
        return jsonify(
            {
                "success": True,
                "saved_value": value,
                "message": "Additional information saved successfully",
            }
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating test procedure: {str(e)}")
        print(f"DEBUG: Error - {str(e)}")
        return jsonify({"success": False, "message": str(e)})


@re_bp.route("/upload_test_procedure_files", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def upload_test_procedure_files():
    try:
        activity_id = request.form.get("activity_id")
        field = request.form.get("field")  # 'walkthrough_files' or 'sampling_files'
        files = request.files.getlist("files")

        print(
            f"DEBUG: Uploading files - activity_id: {activity_id}, field: {field}, files: {[f.filename for f in files]}"
        )

        if not activity_id or not files or files[0].filename == "":
            return jsonify(
                {
                    "success": False,
                    "message": "Missing required parameters or no files selected",
                }
            )

        # Get the project control activity
        project_control = (
            ProjectControlActivity.query.options(
                db.joinedload(ProjectControlActivity.project_test_procedure)
            )
            .filter_by(id=activity_id)
            .first()
        )

        if not project_control:
            return jsonify({"success": False, "message": "Activity not found"})

        # Get or create test procedure
        if not project_control.project_test_procedure:
            new_test_procedure = ProjectTestSteps()
            db.session.add(new_test_procedure)
            project_control.project_test_procedure = new_test_procedure
            db.session.flush()

        # Create upload directory if it doesn't exist
        upload_dir = os.path.join(
            current_app.root_path, "static", "uploads", "test_procedure_files"
        )
        os.makedirs(upload_dir, exist_ok=True)

        uploaded_files = []
        for file in files:
            if file.filename == "":
                continue

            # Generate unique filename
            original_filename = secure_filename(file.filename)
            file_extension = os.path.splitext(original_filename)[1]
            unique_filename = f"{uuid.uuid4()}{file_extension}"
            file_path = os.path.join(upload_dir, unique_filename)

            # Save file
            file.save(file_path)

            # Create file record
            new_file = TestProcedureFile(
                test_procedure_id=project_control.project_test_procedure.id,
                filename=original_filename,
                file_path=unique_filename,
                file_size=os.path.getsize(file_path),
                file_type=file_extension,
                field_type=field,
            )
            db.session.add(new_file)
            uploaded_files.append(original_filename)

        db.session.commit()

        print(f"DEBUG: Successfully uploaded {len(uploaded_files)} files")
        return jsonify(
            {
                "success": True,
                "message": f"Successfully uploaded {len(uploaded_files)} files",
                "files": uploaded_files,
            }
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error uploading test procedure files: {str(e)}")
        print(f"DEBUG: Upload error - {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@re_bp.route("/delete_test_procedure_file", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def delete_test_procedure_file():
    try:
        data = request.get_json()
        file_id = data.get("file_id")
        field = data.get("field")

        print(f"DEBUG: Deleting file - file_id: {file_id}, field: {field}")

        if not file_id:
            return jsonify({"success": False, "message": "File ID is required"})

        # Get file record
        file_record = TestProcedureFile.query.get(file_id)
        if not file_record:
            return jsonify({"success": False, "message": "File not found"})

        # Delete physical file
        file_path = os.path.join(
            current_app.root_path,
            "static",
            "uploads",
            "test_procedure_files",
            file_record.file_path,
        )
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"DEBUG: Deleted physical file: {file_path}")
        else:
            print(f"DEBUG: Physical file not found: {file_path}")

        # Delete database record
        db.session.delete(file_record)
        db.session.commit()

        print("DEBUG: File deleted successfully from database")
        return jsonify(
            {
                "success": True,
                "message": "File deleted successfully",
                "deleted_file_id": file_id,
            }
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting test procedure file: {str(e)}")
        print(f"DEBUG: Delete error - {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@re_bp.route("/get_test_procedure_files", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def get_test_procedure_files():
    try:
        activity_id = request.args.get("activity_id")
        field = request.args.get("field")

        if not activity_id or not field:
            return jsonify({"success": False, "message": "Missing required parameters"})

        project_control = (
            ProjectControlActivity.query.options(
                db.joinedload(ProjectControlActivity.project_test_procedure).joinedload(
                    ProjectTestSteps.test_procedure_files
                )
            )
            .filter_by(id=activity_id)
            .first()
        )

        if not project_control or not project_control.project_test_procedure:
            return jsonify(
                {"success": True, "files": []}
            )  # Return empty array instead of error

        files_data = []
        for file in project_control.project_test_procedure.test_procedure_files:
            if file.field_type == field:
                files_data.append(
                    {
                        "id": file.id,
                        "filename": file.filename,
                        "file_size": file.file_size,
                        "download_url": url_for(
                            "re.download_test_procedure_file",
                            file_id=file.id,
                            _external=True,
                        ),  # Added _external=True
                    }
                )

        return jsonify({"success": True, "files": files_data})

    except Exception as e:
        current_app.logger.error(f"Error fetching test procedure files: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@re_bp.route("/download_test_procedure_file/<int:file_id>")
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def download_test_procedure_file(file_id):
    try:
        file_record = TestProcedureFile.query.get_or_404(file_id)
        file_path = os.path.join(
            current_app.root_path,
            "static",
            "uploads",
            "test_procedure_files",
            file_record.file_path,
        )

        return send_file(
            file_path, as_attachment=True, download_name=file_record.filename
        )
    except Exception as e:
        current_app.logger.error(f"Error downloading file: {str(e)}")
        return "File not found", 404


@re_bp.route("/report", methods=["GET"])
def report():
    return render_template("report.html")


# @re_bp.route('submit_evidance', methods=['POST'])
# def evidances():
#     data = request.get_json()
#     print(data)
#     return jsonify({'message':'Got Input'}), 200

# UPLOAD_FOLDER = "evidences"
# UPLOAD_FOLDER_1 = "uploads\evidences"

# Use pathlib for cross-platform compatibility
UPLOAD_FOLDER = "evidences"
UPLOAD_FOLDER_1 = Path("uploads") / "evidences"  # Use Path for cross-platform

# Make sure the uploads directory exists
UPLOAD_FOLDER_1.mkdir(parents=True, exist_ok=True)

# Convert to string for send_from_directory
UPLOAD_FOLDER_1_STR = str(UPLOAD_FOLDER_1)

ALLOWED_EXTENSIONS = {
    "pdf",
    "png",
    "jpg",
    "jpeg",
    "doc",
    "docx",
    "xls",
    "xlsx",
    "ppt",
    "pptx",
    "csv",
    "txt",
    "mp3",
    "wav",
}


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@re_bp.route("/uploads/evidences/<path:filename>")
def uploaded_file(filename):
    """Serve uploaded evidence files from the uploads/evidences directory."""
    from pathlib import Path

    # Clean the filename
    clean_filename = filename.replace("\\", "/")

    # Remove "evidences/" prefix if it exists
    if clean_filename.startswith("evidences/"):
        clean_filename = clean_filename[10:]  # Remove "evidences/"

    # Security check: prevent directory traversal
    if ".." in clean_filename or clean_filename.startswith("/"):
        abort(404)

    # Use Path for cross-platform compatibility
    upload_dir = Path("uploads") / "evidences"

    # Make it absolute
    if not upload_dir.is_absolute():
        upload_dir = Path.cwd() / upload_dir

    file_path = upload_dir / clean_filename

    print(f"📂 Serving file: {clean_filename}")
    print(f"📁 Upload directory: {upload_dir}")
    print(f"📍 Full path: {file_path}")
    print(f"✅ File exists: {file_path.exists()}")

    if not file_path.exists():
        print(f"❌ File not found at: {file_path}")
        abort(404)

    # Use send_file instead of send_from_directory
    return send_file(
        file_path,
        as_attachment=False,  # Display in browser instead of downloading
        conditional=True,  # Support for conditional requests (ETag, If-Modified-Since)
    )


@re_bp.route("/evidences", methods=["POST"])
def evidences():
    try:
        project_evidence_id = request.form.get("evidence_id", "").strip()
        evidence_item = request.form.get("evidence_item", "").strip()
        evidence_text = request.form.get("evidence_content", "").strip()
        uploaded_path = request.form.get("evidence_uploaded_path", "").strip()
        clause_description = request.form.get("evidence_clause_description")
        file = request.files.get("evidence_file")

        if not project_evidence_id:
            flash("Missing evidence ID.", "warning")
            return redirect(request.referrer)

        artifact = ProjectEvidenceArtifact.query.get(project_evidence_id)
        if not artifact:
            flash(
                f"Project evidence with ID {project_evidence_id} not found.", "danger"
            )
            return redirect(request.referrer)

        if not evidence_text and not file and not uploaded_path:
            flash(
                "Both evidence text and file fields are empty. Nothing to submit.",
                "info",
            )
            return redirect(request.referrer)

        # --- Handle manual text input ---
        if (
            evidence_text
            and evidence_text.strip()
            and evidence_text.strip() != "<p><br></p>"
        ):
            artifact.evidence_text = evidence_text
            flash("Evidence updated with provided text.", "success")
            db.session.commit()
            return redirect(request.referrer)

        # --- Process file for AI autofill ---
        filename = None
        full_physical_file_path = None

        if file and file.filename:
            if allowed_file(file.filename):
                # Generate unique filename
                filename = secure_filename(
                    f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}"
                )
                full_physical_file_path = UPLOAD_FOLDER_1 / filename

                try:
                    file.save(str(full_physical_file_path))
                    print(f"File saved to: {full_physical_file_path}")
                except IOError as io_err:
                    flash(f"Error saving file: {io_err}", "danger")
                    return redirect(request.referrer)
            else:
                flash("Invalid file type.", "warning")
                return redirect(request.referrer)

        elif uploaded_path:
            # Clean up the uploaded path to get just the filename
            # Remove any "evidences/" or "evidences\" prefix
            clean_path = uploaded_path

            # Strip evidences prefix if present
            if clean_path.startswith("evidences/"):
                clean_path = clean_path[10:]  # Remove "evidences/"
            elif clean_path.startswith("evidences\\"):
                clean_path = clean_path[10:]  # Remove "evidences\"

            # If there are still path separators, get the last part
            if "\\" in clean_path:
                filename = clean_path.split("\\")[-1]
            elif "/" in clean_path:
                filename = clean_path.split("/")[-1]
            else:
                filename = clean_path

            full_physical_file_path = UPLOAD_FOLDER_1 / filename

        if filename:
            try:
                # CRITICAL: Store ONLY the filename in database, NO prefix
                artifact.evidence_file_path = filename
                db.session.commit()

                print(f"✅ Stored filename in DB: {filename}")
                print(f"✅ Physical file location: {full_physical_file_path}")
                print(f"✅ File exists: {os.path.exists(full_physical_file_path)}")

                # Only attempt AI processing for supported file types
                file_extensions_to_exclude = (
                    ".xlsx",
                    ".xls",
                    ".csv",
                    ".png",
                    ".jpeg",
                    ".jpg",
                    ".mp3",
                    ".wav",
                )

                if not str(full_physical_file_path).endswith(
                    file_extensions_to_exclude
                ):
                    try:
                        prompt = get_evidence_prompt(
                            project_evidence_id, evidence_item, clause_description
                        )
                        print(prompt)

                        res = extract_structured_info_3(
                            prompt,
                            ComplianceEvidence,
                            str(full_physical_file_path),
                            vector_store_id=None,
                        )
                        print("response", res)

                        if res and hasattr(res, "answer") and res.answer:
                            artifact.evidence_text = res.answer
                            flash(
                                "Evidence updated with AI-generated content!", "success"
                            )
                        else:
                            flash(
                                "File saved but AI could not generate content.",
                                "warning",
                            )

                    except Exception as ai_e:
                        print(f"AI processing error: {ai_e}")
                        flash("File saved but AI processing failed.", "warning")
                else:
                    flash(
                        f"File saved. AI processing not supported for {os.path.splitext(filename)[1]} files.",
                        "info",
                    )

                flash(f"File uploaded successfully!", "success")
                db.session.commit()
                return redirect(request.referrer)

            except Exception as e:
                db.session.rollback()
                print(f"Error processing file: {e}")
                print(f"Traceback: {traceback.format_exc()}")
                flash(f"Error processing file: {e}", "danger")
                return redirect(request.referrer)
        else:
            flash("No content provided to update evidence.", "info")
            return redirect(request.referrer)

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Unexpected error in /evidences route: {e}")
        flash("An unexpected server error occurred.", "danger")
        return redirect(request.referrer)


@re_bp.route("/populate_data", methods=["GET"])
def extract_and_format_data():
    """
    Extracts audit observation data from the database and formats it into
    a list of dictionaries, suitable for document generation.
    Each field's content will be a single string, potentially multi-line,
    formatted under triple quotes for multi-line document insertion.
    """
    try:
        # Template path
        template_docx = "app/routes/re/Fictional audit Template.docx"

        # Query data with necessary joins to prevent N+1 queries
        data = ControlActivity.query.options(
            joinedload(ControlActivity.compliance_activity).joinedload(
                ComplianceActivities.clauses
            ),
            joinedload(ControlActivity.test_procedure).joinedload(TestSteps.documents),
            joinedload(ControlActivity.test_procedure)
            .joinedload(TestSteps.interviews)
            .joinedload(Interview.roles),
            joinedload(ControlActivity.test_procedure)
            .joinedload(TestSteps.interviews)
            .joinedload(Interview.questions),
            joinedload(ControlActivity.evidences),
        ).all()

        print(f"Fetched {len(data)} Control Activities from DB.")

        formatted_observations = []
        for item in data:
            d1 = {}

            # Reference no.
            d1["Reference no."] = (
                item.compliance_activity.clauses.clause_no
                if item.compliance_activity and item.compliance_activity.clauses
                else "N/A"
            )

            # Guideline Name
            try:
                guideline_data = (
                    item.compliance_activity.clauses.guideline.guideline_data
                )
                d1["Guideline Name"] = guideline_data.get("DocumentDetails", {}).get(
                    "DocumentName", "N/A"
                )
            except Exception:
                d1["Guideline Name"] = "N/A"

            # Control Objective
            d1["Control Objective"] = getattr(item, "objective", "N/A")

            # Test to be performed
            test_procedure_parts = []
            tp = item.test_procedure
            if tp:
                test_procedure_parts.append(
                    f"Walkthrough:\n  {getattr(tp, 'walkthrough', 'N/A')}"
                )
                test_procedure_parts.append(
                    f"Sampling:\n  {getattr(tp, 'sampling', 'N/A')}"
                )

                if tp.documents:
                    doc_names = [
                        doc.document_name for doc in tp.documents if doc.document_name
                    ]
                    test_procedure_parts.append(
                        f"Documents Reviewed:\n  {', '.join(doc_names) if doc_names else 'N/A'}"
                    )

                interview = tp.interviews
                if interview:
                    if interview.roles:
                        roles = [r.role for r in interview.roles if r.role]
                        test_procedure_parts.append(
                            f"Interview Roles:\n  {', '.join(roles)}"
                        )
                    if interview.questions:
                        q_a_pairs = [
                            f"Q: {q.question}\n"
                            for q in interview.questions
                            if q.question
                        ]
                        test_procedure_parts.append(
                            "Interview Questions:\n" + "\n".join(q_a_pairs)
                        )

            d1["Test to be performed"] = (
                "\n\n".join(test_procedure_parts) if test_procedure_parts else "N/A"
            )

            # Detailed observation
            d1["Detailed observation"] = getattr(item, "auditor_observation", "N/A")

            # Gap noted & Impact
            gap_impact_parts = []
            if item.findings:
                gap_impact_parts.append(f"Gap: {item.findings}")
            if item.impact:
                gap_impact_parts.append(f"Impact: {item.impact}")
            d1["Gap noted & Impact"] = (
                "\n".join(gap_impact_parts) if gap_impact_parts else "N/A"
            )

            # Severity
            d1["Severity"] = getattr(item, "severity", "N/A")

            # Recommendation
            d1["Recommendation"] = getattr(item, "recommendations", "N/A")

            # Evidences
            evidences_formatted_parts = []
            for evidence in item.evidences:
                parts = []
                if evidence.item:
                    parts.append(f"Item: {evidence.item}")
                if evidence.evidance:
                    parts.append(
                        f"Description: {clean_html_preserve_format(evidence.evidance)}"
                    )
                if evidence.evidance_file:
                    parts.append(
                        f"File: {url_for('uploaded_file', filename=evidence.evidance_file, _external=True)}"
                    )
                evidences_formatted_parts.append(" - ".join(parts))

            d1["References to evidences/ Proof of concept"] = (
                "\n\n".join(evidences_formatted_parts)
                if evidences_formatted_parts
                else "N/A"
            )

            formatted_observations.append(d1)

        print("Formatted Data (first 2 entries for brevity):")
        for obs in formatted_observations[:2]:
            print(obs)
        if len(formatted_observations) > 2:
            print(f"... and {len(formatted_observations) - 2} more observations.")

        # Output doc path
        output_doc_path = (
            "uploads/audit_report/Updated_Audit_Report_Separate_Tables.docx"
        )
        os.makedirs(os.path.dirname(output_doc_path), exist_ok=True)

        process_audit_observations(
            template_docx, output_doc_path, formatted_observations
        )

        return jsonify(formatted_observations), 200

    except Exception as e:
        print(f"An error occurred during data extraction: {e}")
        # Log the full traceback for detailed error analysis in a real application
        import traceback

        traceback.print_exc()
        return jsonify({"error": f"Failed to extract and format data: {e}"}), 500


@re_bp.route("/report_options/<int:project_id>", methods=["GET"])
def report_options(project_id):
    """
    Show report generation options page
    """
    try:
        project = Projects.query.get(project_id)

        if not project:
            flash("Project not found", "error")
            return redirect(url_for("audit.my_projects"))

        free_report_used = current_user.free_report_used if current_user else False

        return render_template(
            "dashboards/auditor/report_options.html",
            project=project,
            current_user=current_user,
            free_report_used=free_report_used,
            project_id=project.id,
        )

    except Exception as e:
        print(f"Error loading report options: {e}")
        flash("Error loading report options", "error")
        return redirect(url_for("audit.my_projects"))


@re_bp.route("/generate_report/<int:project_id>", methods=["POST"])
def generate_audit_report(project_id):
    """
    Generate audit report using ONLY consolidated clause data
    """
    try:

        # Check if free report was already used by the user
        if current_user.free_report_used:
            flash(
                "You have already used your free report. Please contact CompliFyre@crackerjacktech.com for additional reports.",
                "warning",
            )
            return redirect(url_for("re.report_options", project_id=project_id))

        # Check if report was already generated
        project = Projects.query.get(project_id)

        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Get form data
        report_format = request.form.get("report_format", "sectional")
        details_include = request.form.getlist("details_include")
        doc_sections = request.form.getlist("doc_sections")
        clause_selection = request.form.get("clause_selection", "all_clauses")

        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Get documentation data if any sections are selected
        documentation_data = {}
        if doc_sections:
            documentation = Documentation.query.filter_by(
                project_id=project.id, created_by=current_user.id
            ).first()

            if documentation:

                # Get clause statistics, severity stats, evidence stats for executive summary
                clause_statistics = get_project_clause_statistics(project_id)
                severity_stats = get_project_severity_statistics(project_id)
                evidence_stats = get_project_evidence_statistics(project_id)

                documentation_data = {
                    "document_control": (
                        get_document_control_data(documentation)
                        if "document_control" in doc_sections
                        else None
                    ),
                    "introduction": (
                        documentation.introduction
                        if "introduction" in doc_sections
                        else None
                    ),
                    "engagement_scope": (
                        documentation.engagement_scope
                        if "engagement_scope" in doc_sections
                        else None
                    ),
                    "auditing_team": (
                        get_auditing_team_data(documentation)
                        if "auditing_team" in doc_sections
                        else None
                    ),
                    "activities_timelines": (
                        documentation.activities_timelines
                        if "activities_timelines" in doc_sections
                        else None
                    ),
                    "methodology_criteria": (
                        documentation.methodology_criteria
                        if "methodology_criteria" in doc_sections
                        else None
                    ),
                    "tools_software": (
                        get_tools_software_data(documentation)
                        if "tools_software" in doc_sections
                        else None
                    ),

                    # NEW: Add executive summary data
                    "executive_summary": (
                        get_executive_summary_data(
                            documentation, 
                            project, 
                            clause_statistics, 
                            severity_stats, 
                            evidence_stats
                        ) if "executive_summary" in doc_sections
                        else None
                    ),
                }

        # Get consolidated evidence for the project
        consolidated_evidence_record = ConsolidatedEvidence.query.filter_by(
            project_id=str(project.id)  # Convert to string
        ).first()

        consolidated_evidence_data = {}
        if (
            consolidated_evidence_record
            and consolidated_evidence_record.consolidate_evidence
        ):
            consolidated_evidence_data = (
                consolidated_evidence_record.consolidate_evidence
            )
            print(
                f"DEBUG: Found consolidated evidence with {len(consolidated_evidence_data.get('grouped_evidences', []))} evidence groups"
            )

            # DEBUGGING: Check the actual data structure
            print(
                f"DEBUG: Consolidated evidence data structure: {list(consolidated_evidence_data.keys())}"
            )
            if "grouped_evidences" in consolidated_evidence_data:
                print(
                    f"DEBUG: Number of evidence groups: {len(consolidated_evidence_data['grouped_evidences'])}"
                )
                for i, group in enumerate(
                    consolidated_evidence_data["grouped_evidences"]
                ):
                    print(f"DEBUG: Group {i}: {group.get('evidence_item_name')}")
                    print(
                        f"DEBUG:   Required by clauses: {group.get('required_by', {}).get('clause_nos', [])}"
                    )
                    print(f"DEBUG:   Evidence items: {len(group.get('evidence', []))}")
            else:
                print(
                    "DEBUG: No 'grouped_evidences' key found in consolidated evidence data"
                )
        else:
            print("DEBUG: No consolidated evidence found for project")

        # Get all applicable clauses for the project
        clauses_query = (
            db.session.query(ProjectClause)
            .join(ProjectGuideline)
            .filter(ProjectGuideline.project_id == project.id)
            .filter(ProjectClause.applicability == True)
            .options(
                joinedload(ProjectClause.project_guideline)
            )  # Eager load guideline data
        )

        # Execute the query
        clauses = clauses_query.all()

        # Natural sorting function - MUST BE THE SAME AS IN ACTIVITY ROUTE
        def natural_sort_key(s):
            """Natural sorting function - identical to activity route"""
            if s is None or s == "":
                return [float("inf")]
            return [
                int(part) if part.isdigit() else part.lower()
                for part in re.split(r"(\d+)", str(s))
            ]

        # Sort clauses using the SAME natural sorting logic
        clauses = sorted(clauses, key=lambda c: natural_sort_key(c.clause_no))

        # Optional: Debug logging to verify sorting
        current_app.logger.info(
            f"Sorted clause order for report {project.project_name}:"
        )
        for i, clause in enumerate(clauses):
            current_app.logger.info(
                f"  {i+1}. Clause {clause.clause_no}: {clause.clause_text[:50]}..."
            )

        consolidated_observations = []

        for clause in clauses:
            # Get clause status info
            clause_status_info = calculate_clause_compliance_status(clause.id)

            # Skip non-compliant clauses if only findings are requested
            if clause_selection == "findings_only":
                if clause_status_info["text"] in ["Compliant", "Fully Compliant"]:
                    continue

            # Get evidence data for this clause using the same logic as clause_test_steps
            clause_evidences = get_evidences_for_clause(clause.id)

            # Get applicability, assessment status, and severity for this clause
            clause_applicability = get_clause_applicability_status(clause.id)
            clause_assessment_status = get_clause_assessment_status(clause.id)
            clause_severity = get_clause_severity(clause.id)
            # Get evidence availability for this clause (Yes if 100% activities have evidence)
            evidence_available = get_clause_evidence_availability(clause.id)

            # Get consolidated summaries
            consolidated_summary_record = ClauseConsolidatedSummary.query.filter_by(
                clause_id=clause.id
            ).first()

            consolidated_test_record = (
                ConsolidatedTestSummary.query.filter_by(clause_id=clause.id)
                .order_by(ConsolidatedTestSummary.generated_at.desc())
                .first()
            )

            consolidated_observation_record = (
                ConsolidatedObservationSummary.query.filter_by(clause_id=clause.id)
                .order_by(ConsolidatedObservationSummary.generated_at.desc())
                .first()
            )

            # Get findings and recommendations from the new consolidated tables
            consolidated_findings_record = (
                ConsolidatedFindingsSummary.query.filter_by(clause_id=clause.id)
                .order_by(ConsolidatedFindingsSummary.updated_at.desc())
                .first()
            )

            consolidated_recommendations_record = (
                ConsolidatedRecommendationsSummary.query.filter_by(clause_id=clause.id)
                .order_by(ConsolidatedRecommendationsSummary.updated_at.desc())
                .first()
            )

            # Format observation data for reporting
            observation_data = {}

            # NEW FIELDS: Add applicability, assessment status, and severity
            observation_data["Applicability Status"] = clause_applicability
            observation_data["Assessment Status"] = clause_assessment_status
            observation_data["Evidence Available"] = evidence_available  
            observation_data["Severity"] = clause_severity

            # Reference no.
            observation_data["Reference no."] = clause.clause_no

            # Add Clause Description/Text
            observation_data["Clause Description"] = clause.clause_text

            # Guideline Name - Get actual guideline name from ProjectGuideline
            guideline_extract = get_guideline_extract(clause)
            observation_data["Guideline Name"] = guideline_extract

            # Test to be performed (Consolidated Test Procedure)
            test_procedure_text = "N/A"
            if consolidated_test_record:
                try:
                    test_summary = json.loads(
                        consolidated_test_record.consolidated_summary
                    )
                    test_parts = []

                    if test_summary.get("consolidated_summary"):
                        test_parts.append(
                            f"Summary: {test_summary['consolidated_summary']}"
                        )

                    if test_summary.get("key_testing_areas"):
                        test_parts.append(
                            f"Key Testing Areas:\n- "
                            + "\n- ".join(test_summary["key_testing_areas"])
                        )

                    if test_summary.get("walkthrough_approach"):
                        test_parts.append(
                            f"Walkthrough Approach: {test_summary['walkthrough_approach']}"
                        )

                    if test_summary.get("sampling_methodology"):
                        test_parts.append(
                            f"Sampling Methodology: {test_summary['sampling_methodology']}"
                        )

                    test_procedure_text = "\n\n".join(test_parts)
                except (json.JSONDecodeError, TypeError):
                    test_procedure_text = consolidated_test_record.consolidated_summary

            observation_data["Test to be performed"] = test_procedure_text

            # Detailed observation (Consolidated Observation Summary)
            observation_text = "N/A"
            if consolidated_observation_record:
                try:
                    obs_summary = json.loads(
                        consolidated_observation_record.consolidated_observation
                    )
                    obs_parts = []

                    if obs_summary.get("consolidated_summary"):
                        obs_parts.append(
                            f"Summary: {obs_summary['consolidated_summary']}"
                        )

                    if obs_summary.get("key_observations"):
                        obs_parts.append(
                            f"Key Observations:\n- "
                            + "\n- ".join(obs_summary["key_observations"])
                        )

                    if obs_summary.get("common_patterns"):
                        obs_parts.append(
                            f"Common Patterns:\n- "
                            + "\n- ".join(obs_summary["common_patterns"])
                        )

                    if obs_summary.get("risk_areas"):
                        obs_parts.append(
                            f"Risk Areas:\n- " + "\n- ".join(obs_summary["risk_areas"])
                        )

                    if obs_summary.get("improvement_opportunities"):
                        obs_parts.append(
                            f"Improvement Opportunities:\n- "
                            + "\n- ".join(obs_summary["improvement_opportunities"])
                        )

                    observation_text = "\n\n".join(obs_parts)
                except (json.JSONDecodeError, TypeError):
                    observation_text = (
                        consolidated_observation_record.consolidated_observation
                    )

            observation_data["Detailed observation"] = observation_text

            # Gap noted & Impact (Consolidated Findings)
            findings_text = "N/A"
            if (
                consolidated_findings_record
                and consolidated_findings_record.consolidated_findings
            ):
                try:
                    findings_data = json.loads(
                        consolidated_findings_record.consolidated_findings
                    )

                    # Extract findings from the new structure
                    if findings_data.get("consolidated_summary"):
                        findings_list = findings_data["consolidated_summary"]
                        if isinstance(findings_list, list) and findings_list:
                            findings_text = "Consolidated Findings:\n• " + "\n• ".join(
                                findings_list
                            )
                        else:
                            findings_text = "Consolidated Findings:\n• " + str(
                                findings_list
                            )
                    else:
                        # Fallback: try to extract from any available structure
                        findings_text = str(findings_data)

                except (json.JSONDecodeError, TypeError, AttributeError) as e:
                    print(f"Error parsing findings for clause {clause.id}: {e}")
                    # Fallback: use raw text
                    findings_text = consolidated_findings_record.consolidated_findings

            observation_data["Gap noted & Impact"] = findings_text

            # Recommendation (Consolidated Recommendations)
            recommendations_text = "N/A"
            if (
                consolidated_recommendations_record
                and consolidated_recommendations_record.consolidated_recommendations
            ):
                try:
                    recommendations_data = json.loads(
                        consolidated_recommendations_record.consolidated_recommendations
                    )

                    # Extract recommendations from the new structure
                    if recommendations_data.get("consolidated_summary"):
                        rec_list = recommendations_data["consolidated_summary"]
                        if isinstance(rec_list, list) and rec_list:
                            recommendations_text = (
                                "Consolidated Recommendations:\n• "
                                + "\n• ".join(rec_list)
                            )
                        else:
                            recommendations_text = (
                                "Consolidated Recommendations:\n• " + str(rec_list)
                            )
                    else:
                        # Fallback: try to extract from any available structure
                        recommendations_text = str(recommendations_data)

                except (json.JSONDecodeError, TypeError, AttributeError) as e:
                    print(f"Error parsing recommendations for clause {clause.id}: {e}")
                    # Fallback: use raw text
                    recommendations_text = (
                        consolidated_recommendations_record.consolidated_recommendations
                    )

            observation_data["Recommendation"] = recommendations_text

            # References to evidences - Use evidence data from clause_test_steps
            evidence_text = format_evidences_for_report(clause_evidences)
            observation_data["References to evidences/ Proof of concept"] = (
                evidence_text
            )

            # Compliance Status
            observation_data["Compliance Status"] = clause_status_info["text"]

            consolidated_observations.append(observation_data)

        # Create a mapping of clause IDs to their activities
        clause_activities = {}
        for clause in clauses:
            clause_activities[clause.id] = get_activities_for_clause_table(clause.id)
        
        # Store this in the report data for later use
        activities_table_data = {
            'clause_activities': clause_activities
        }    

        BASE_DIR = os.path.abspath(
            os.path.join(
                os.path.dirname(__file__),
                "../../../",
            )
        )

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        # After successfully generating the report, update project status
        project.report_generated = True
        project.report_generated_at = datetime.now()
        project.report_generated_by = current_user.id
        project.free_report_used = True  # Mark free report as used

        current_user.free_report_used = True

        db.session.commit()

        if report_format == "sectional":
            # Generate Word document
            template_docx = "app/routes/re/Fictional audit Template.docx"
            output_doc_path = f"uploads/audit_report/{project.project_name}_Consolidated_Audit_Report.docx"
            full_path = os.path.join(BASE_DIR, output_doc_path)
            os.makedirs(os.path.dirname(full_path), exist_ok=True)

            process_audit_observations(
                template_docx, full_path, consolidated_observations, documentation_data,clauses=clauses,  clause_activities=clause_activities  # Pass the activities data
            )

            return send_file(full_path, as_attachment=True)

        else:
            # Generate Excel file
            output_excel_path = f"uploads/audit_report/{project.project_name}_Consolidated_Audit_Report.xlsx"
            full_path = os.path.join(BASE_DIR, output_excel_path)
            os.makedirs(os.path.dirname(full_path), exist_ok=True)

            # Create Excel workbook
            wb = Workbook()

            # Remove the default sheet if we're adding documentation
            if doc_sections and documentation_data:
                wb.remove(wb.active)

            # Add documentation sections as separate sheets if selected
            if doc_sections and documentation_data:
                print(f"Adding documentation sheets: {doc_sections}")
                add_documentation_sheets_to_excel(wb, documentation_data)
            else:
                print("No documentation sections selected or no documentation data")

            # Create main observations sheet
            ws = wb.create_sheet("Consolidated Audit Observations")

            # If no documentation sheets were added, we need at least one sheet
            if len(wb.sheetnames) == 0:
                ws = wb.active
                ws.title = "Consolidated Audit Observations"

            # Add headers for consolidated observations
            if consolidated_observations:
                headers = list(consolidated_observations[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    # Style header
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).fill = PatternFill(
                        start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
                    )

                # Add data
                for row, observation in enumerate(consolidated_observations, 2):
                    for col, key in enumerate(headers, 1):
                        value = observation.get(key, "N/A")
                        if isinstance(value, str):
                            value = value.replace("\n\n", "\n").strip()
                        ws.cell(row=row, column=col, value=value)

                # Auto-adjust column widths for all sheets
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(full_path)
                print(
                    f"Excel file saved with {len(wb.sheetnames)} sheets: {wb.sheetnames}"
                )
                return send_file(full_path, as_attachment=True)

    except Exception as e:
        print(f"An error occurred during consolidated report generation: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Failed to generate consolidated report: {e}", "error")
        return redirect(url_for("re.report_options", project_id=project_id))


def check_report_generated(project_id):
    """Utility function to check if report was generated for a project"""
    project = Projects.query.get(project_id)
    if project and project.report_generated:
        flash(
            "Editing is not allowed after report generation. Please contact CompliFyre@crackerjacktech.com if you need to make changes.",
            "error",
        )
        return True
    return False


def get_evidences_for_clause(clause_id):
    """
    Get all evidences for a clause using the same logic as clause_test_steps
    """
    all_evidences = []

    # Get only APPLICABLE project compliance activities for this clause
    project_compliance_activities = (
        db.session.query(ProjectComplianceActivity)
        .filter_by(project_clause_id=clause_id, applicability=True)
        .all()
    )

    for pca in project_compliance_activities:
        # Get all control activities for this compliance activity
        control_activities = (
            db.session.query(ProjectControlActivity)
            .filter_by(project_compliance_activity_id=pca.id)
            .options(
                joinedload(ProjectControlActivity.submitted_evidences),
                joinedload(ProjectControlActivity.project_compliance_activity),
            )
            .all()
        )

        for control_activity in control_activities:
            display_activity_name = (
                control_activity.project_compliance_activity.activity_description
            )

            # Add activity information with evidences
            for evidence in control_activity.submitted_evidences:
                all_evidences.append(
                    {
                        "activity_name": display_activity_name,
                        "activity_code": control_activity.activity_code,
                        "category": evidence.category,
                        "item": evidence.item,
                        "evidence_text": evidence.evidence_text,
                        "evidence_file_path": evidence.evidence_file_path,
                        "id": evidence.id,
                        "is_applicable": True,
                    }
                )

    print(f"DEBUG: Found {len(all_evidences)} evidences for clause {clause_id}")
    return all_evidences


def get_guideline_extract(clause):
    """
    Extract guideline information from ProjectGuideline for a given clause
    """
    try:
        if clause.project_guideline and clause.project_guideline.guideline_data:
            guideline_data = clause.project_guideline.guideline_data

            # Extract guideline name from the JSON structure
            if isinstance(guideline_data, dict):
                # Try different possible keys for guideline name
                guideline_name = (
                    guideline_data.get("DocumentDetails", {}).get("DocumentName")
                    or guideline_data.get("document_name")
                    or guideline_data.get("name")
                    or guideline_data.get("title")
                    or "Unnamed Guideline"
                )

                # Also try to get document ID if available
                document_id = (
                    guideline_data.get("DocumentDetails", {}).get("DocumentID")
                    or guideline_data.get("document_id")
                    or guideline_data.get("id")
                    or ""
                )

                # Format the Guideline Name
                if document_id:
                    return f"{guideline_name} (ID: {document_id}) - Clause {clause.clause_no}"
                else:
                    return f"{guideline_name} - Clause {clause.clause_no}"

            # If guideline_data is not a dict, try to use it as string
            elif isinstance(guideline_data, str):
                try:
                    # Try to parse as JSON
                    parsed_data = json.loads(guideline_data)
                    if isinstance(parsed_data, dict):
                        guideline_name = (
                            parsed_data.get("DocumentDetails", {}).get("DocumentName")
                            or parsed_data.get("document_name")
                            or parsed_data.get("name")
                            or parsed_data.get("title")
                            or "Unnamed Guideline"
                        )
                        return f"{guideline_name} - Clause {clause.clause_no}"
                except (json.JSONDecodeError, TypeError):
                    # If it's not JSON, use the string directly
                    return f"{guideline_data} - Clause {clause.clause_no}"

        # Fallback if no guideline data is found
        return f"Guideline - Clause {clause.clause_no}: {clause.clause_text}"

    except Exception as e:
        print(f"Error extracting guideline for clause {clause.clause_no}: {e}")
        return f"Guideline - Clause {clause.clause_no}: {clause.clause_text}"


def format_evidences_for_report(evidences):
    """
    Format evidence data for report display
    """
    if not evidences:
        return "No evidence requirements specified for this clause"

    # Group by activity
    activity_evidences = {}
    for evidence in evidences:
        activity_key = f"{evidence['activity_code']} - {evidence['activity_name']}"
        if activity_key not in activity_evidences:
            activity_evidences[activity_key] = []

        if evidence["item"] and evidence["item"].strip():
            activity_evidences[activity_key].append(evidence["item"].strip())

    # Format the output
    evidence_sections = []

    for activity, items in activity_evidences.items():
        if items:
            section = f"{activity}:\n" + "\n".join([f"  • {item}" for item in items])
            evidence_sections.append(section)

    if evidence_sections:
        return "Evidence Requirements:\n\n" + "\n\n".join(evidence_sections)
    else:
        return "No evidence items found for this clause"


@re_bp.route("/activity_by_guideline/<int:id>", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def activity_by_guideline(id):
    """
    Show all ComplianceActivities linked to a given Guideline ID
    """
    try:
        add_to_breadcrumb(request.full_path, "Activity for Guidelines")
        # Step 1: Get all Clauses for the given Guideline

        clauses = Clauses.query.filter_by(guideline_id=id).all()
        guidelines = Guidelines.query.all()
        clause_ids = [clause.id for clause in clauses]
        project_names = db.session.query(Projects.project_name).distinct().all()

        if not clause_ids:
            return render_template("compliance_activity.html", compliance_activities=[])

        # Step 2: Get all ComplianceActivities linked to those Clauses
        compliance_activities = ComplianceActivities.query.filter(
            ComplianceActivities.clause_id.in_(clause_ids)
        ).all()

        # Step 3: Pass to template
        return render_template(
            "compliance_activity.html",
            compliance_activities=compliance_activities,
            guidelines=guidelines,
            project=project_names,
        )

    except Exception as err:
        current_app.logger.error(f"Unexpected error: {str(err)}")
        return jsonify({"error": "Internal server error"}), 500


@re_bp.route("/test_procedure_artifacts/<int:activity_id>")
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def test_procedure_artifacts(activity_id):
    try:
        add_to_breadcrumb(request.full_path, "Master Test Procedure")
        control = (
            db.session.query(ControlActivity)
            .filter_by(compliance_activity_id=activity_id)
            .options(
                db.joinedload(ControlActivity.test_procedure).joinedload(
                    TestSteps.documents
                ),
                db.joinedload(ControlActivity.test_procedure)
                .joinedload(TestSteps.interviews)
                .joinedload(Interview.roles),
                db.joinedload(ControlActivity.test_procedure)
                .joinedload(TestSteps.interviews)
                .joinedload(Interview.questions),
                db.joinedload(ControlActivity.evidences),
            )
            .first()
        )

        if not control:
            return render_template("test_artifacts.html", how_to_perform_activities=[])

        test_procedure = control.test_procedure
        interview_obj = test_procedure.interviews if test_procedure else None

        document_reviews = (
            [doc.document_name for doc in test_procedure.documents]
            if test_procedure
            else []
        )

        interview_data = {
            "roles": (
                [role.role for role in interview_obj.roles] if interview_obj else []
            ),
            "key_questions": (
                [
                    {"id": q.id, "question": q.question, "answer": q.answer}
                    for q in interview_obj.questions
                ]
                if interview_obj
                else []
            ),
        }

        evidence_data = {}
        for evidence in control.evidences:
            evidence_entry = {
                "id": evidence.id,
                "item": evidence.item,
                "evidence_ans": evidence.evidance,
                "files": evidence.evidance_file,
            }
            evidence_data.setdefault(evidence.category, []).append(evidence_entry)

        activity_data = {
            "activity_id": activity_id,
            "activity_code": control.activity_code,
            "activity_name": control.activity_name,
            "activity_description": control.activity_description,
            "objective": control.objective,
            "owner": control.owner,
            "control_type": control.control_type,
            "frequency": control.frequency,
            "sampling_guidance": control.sampling_guidance,
            "auditor_observation": control.auditor_observation,
            "findings": control.findings,
            "impact": control.impact,
            "severity": control.severity,
            "recommendations": control.recommendations,
            "reviewer_notes": control.reviewer_notes,
            "explain_test_procedure": control.explain_test_procedure,
            "test_procedure": {
                "Walkthrough": test_procedure.walkthrough if test_procedure else None,
                "sampling": test_procedure.sampling if test_procedure else None,
                "review_of_documentation": document_reviews,
                "interviews": interview_data,
            },
            "evidences_artifacts_needed": evidence_data,
            "compliant_status": control.compliant_status,
            "control_finding": control.control_findings,
            "control_recomend": control.control_recommendation,
        }

        # Pass a list with one item to keep same data shape as before
        return render_template(
            "master_test_procedure.html",
            how_to_perform_activities=[{"data": activity_data}],
        )

    except Exception as e:
        current_app.logger.error(f"Unexpected error: {str(e)}")
        return render_template(
            "master_test_procedure.html", how_to_perform_activities=[]
        )


@re_bp.route("/applicability_status", methods=["GET"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def applicability_status():
    activity_id = request.args.get("activity_id")

    if not activity_id:
        return {"status": "error", "message": "No activity ID provided."}, 400

    activity = ProjectComplianceActivity.query.get(activity_id)
    if not activity:
        return {"status": "error", "message": "Activity not found."}, 404

    try:
        activity.applicability = not activity.applicability
        db.session.commit()
        return {
            "status": "success",
            "message": f"Applicability status changed to {'Applicable' if activity.applicability else 'Not Applicable'}.",
            "applicability": activity.applicability,
        }, 200
    except Exception as e:
        db.session.rollback()
        return {"status": "error", "message": f"Database update failed: {str(e)}"}, 500


@re_bp.route("/clause_applicability_status", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def clause_applicability_status():
    """
    Toggle or set applicability for a ProjectClause.
    Simple and reliable like the activity applicability.
    """
    try:
        # Get data from request
        data = {}

        if request.is_json:
            data = request.get_json(silent=True) or {}
        else:
            # Handle form data
            data = request.form.to_dict()

        clause_id = data.get("clause_id")

        if not clause_id:
            return jsonify(status="error", message="Missing clause_id"), 400

        # Convert to int
        try:
            clause_id = int(clause_id)
        except (ValueError, TypeError):
            return jsonify(status="error", message="Invalid clause_id"), 400

        clause = ProjectClause.query.get(clause_id)
        if not clause:
            return jsonify(status="error", message="Clause not found"), 404

        # Get new applicability value
        # If provided, use it; otherwise toggle
        if "applicability" in data:
            # Handle different formats
            applicability_str = str(data.get("applicability")).lower()
            if applicability_str in ("true", "1", "yes", "on"):
                new_applicability = True
            elif applicability_str in ("false", "0", "no", "off"):
                new_applicability = False
            else:
                # If invalid, toggle
                new_applicability = not clause.applicability
        else:
            # Toggle if not specified
            new_applicability = not clause.applicability

        # Update the clause
        clause.applicability = new_applicability
        db.session.commit()

        # Return simple success response
        return (
            jsonify(
                status="success",
                applicability=new_applicability,
                clause_id=clause.id,
                message=f"Clause applicability updated to {'Applicable' if new_applicability else 'Not Applicable'}",
            ),
            200,
        )

    except Exception as e:
        current_app.logger.error(f"Error updating clause applicability: {str(e)}")
        db.session.rollback()
        return jsonify(status="error", message=f"Update failed: {str(e)}"), 500


@re_bp.route("/delete_test_procedure/<int:id>", methods=["GET"])
def delete_test_procedure(id):
    """
    Deletes a specific compliance activity by its ID.
    The `cascade="all, delete-orphan"` on the relationships in the model will
    automatically delete associated records in other tables.
    """
    # Query the database to find the compliance activity with the given ID.
    activity_to_delete = db.session.get(ControlActivity, id)

    # Check if the activity was found.
    if activity_to_delete:
        try:
            # Delete the activity from the database session.
            db.session.delete(activity_to_delete)
            # Commit the changes to the database.
            db.session.commit()
            flash(f"Activity Deleted Successfully", "success")
            return redirect(request.referrer)
        except Exception as e:
            # Rollback the transaction in case of an error.
            db.session.rollback()
            flash("Something Went Wrong", "error")
            return redirect(request.referrer)
    else:
        # If the activity was not found, return a 404 Not Found error.
        flash("Activity not found,", "error")
        return redirect(request.referrer)


@re_bp.route("/control-activities/<int:control_id>", methods=["GET"])
def update_control_activity(control_id: int):
    """
    Updates a ControlActivity and its related records using a new LLM-generated payload.
    It fetches the existing control activity, gets its description and associated clause
    text, and then simulates a call to an LLM to get updated data.
    """
    try:
        # data = request.get_json()
        # vec_id = data.get('vec_id')
        # if not vec_id:
        #     return jsonify({"error": "Missing 'vec_id' in request body"}), 400

        with session_scope() as session:
            # Eagerly load the related data to avoid multiple queries later
            control = (
                session.query(ControlActivity)
                .options(
                    joinedload(ControlActivity.compliance_activity).joinedload(
                        ComplianceActivities.clauses
                    )
                )
                .filter(ControlActivity.id == control_id)
                .first()
            )

            if not control:
                return (
                    jsonify(
                        {"error": f"ControlActivity with ID {control_id} not found"}
                    ),
                    404,
                )

            # Get the required data from the existing objects
            clause_text = control.compliance_activity.clauses.clause_text
            compliance_activity_payload = {
                "activity_description": control.activity_description
            }

            # Simulate LLM call to get updated data
            updated_data = extract_structured_info_2(
                test_procedure(clause_text, compliance_activity_payload),
                ControlWorkpaper,
            )
            updated_data_dict = _as_dict(updated_data) or {}

            # --- Update the ControlActivity record itself ---
            control.activity_code = updated_data_dict.get("activity_code")
            control.activity_name = updated_data_dict.get("activity_name")
            control.activity_description = updated_data_dict.get("activity_description")
            control.objective = updated_data_dict.get("objective")
            control.owner = updated_data_dict.get("owner")
            control.control_type = updated_data_dict.get("control_type")
            control.frequency = updated_data_dict.get("frequency")
            control.sampling_guidance = updated_data_dict.get("sampling_guidance")
            control.auditor_observation = updated_data_dict.get("auditor_observation")
            control.findings = updated_data_dict.get("findings")
            control.impact = updated_data_dict.get("impact")
            control.severity = updated_data_dict.get("severity")
            control.recommendations = updated_data_dict.get("recommendations")
            control.reviewer_notes = updated_data_dict.get("reviewer_notes")
            control.explain_test_procedure = updated_data_dict.get(
                "explain_test_procedure"
            )

            # --- Handle related records (TestSteps, Interviews, etc.) ---
            # SQLAlchemy's cascade="all, delete-orphan" makes this easy.
            # We first clear the relationship, which triggers deletion of old children.
            if control.test_procedure:
                session.delete(control.test_procedure)
                control.test_procedure = None

            # Now, add the new TestSteps and its children
            test_steps_payload = updated_data_dict.get("test_procedure") or {}
            new_test_steps = TestSteps(
                walkthrough=_ci_get(test_steps_payload, "walkthrough"),
                sampling=_ci_get(test_steps_payload, "sampling"),
                control_id=control.id,
            )
            session.add(new_test_steps)
            session.flush()

            # Document reviews
            docs_list = (
                test_steps_payload.get("review_of_documentation")
                or test_steps_payload.get("review_of_documents")
                or []
            )
            if isinstance(docs_list, (str, bytes)):
                docs_list = [docs_list]
            for doc in docs_list:
                session.add(
                    DocumentReview(
                        test_procedure_id=new_test_steps.id, document_name=doc
                    )
                )

            # Interviews
            interviews_data = test_steps_payload.get("interviews") or {}
            new_interview = Interview(test_procedure_id=new_test_steps.id)
            session.add(new_interview)
            session.flush()
            for role in interviews_data.get("roles") or []:
                session.add(InterviewRole(interview_id=new_interview.id, role=role))
            for question in (
                interviews_data.get("key_questions")
                or interviews_data.get("questions")
                or []
            ):
                session.add(
                    InterviewQuestion(interview_id=new_interview.id, question=question)
                )

            # Evidence artifacts: Clear existing and add new
            control.evidences.clear()
            evidence_input = updated_data_dict.get("evidences_artifacts_needed") or []
            if isinstance(evidence_input, dict):
                iter_evidence = (
                    (k, v if isinstance(v, list) else [v])
                    for k, v in evidence_input.items()
                )
            elif isinstance(evidence_input, list):

                def _yield_from_list(lst):
                    for entry in lst:
                        if not entry:
                            continue
                        if isinstance(entry, dict):
                            cat = (
                                entry.get("category") or entry.get("name") or "Unknown"
                            )
                            items = entry.get("items") or entry.get("items_list") or []
                            yield (cat, items if isinstance(items, list) else [items])
                        else:
                            yield ("Unknown", [str(entry)])

                iter_evidence = _yield_from_list(evidence_input)
            else:
                iter_evidence = []

            for category, items in iter_evidence:
                category = (category or "Unknown").strip()
                for item in items or []:
                    artifact = (
                        session.query(EvidenceArtifact)
                        .filter_by(category=category, item=item)
                        .first()
                    )
                    if not artifact:
                        artifact = EvidenceArtifact(category=category, item=item)
                        session.add(artifact)
                    control.evidences.append(artifact)

            session.flush()
            session.commit()
            logger.info("ControlActivity with ID %s updated successfully.", control_id)
            return (
                jsonify(
                    {
                        "message": f"ControlActivity with ID {control_id} updated successfully."
                    }
                ),
                200,
            )

    except Exception as e:
        logger.error(
            "Error updating ControlActivity with ID %s: %s", control_id, str(e)
        )
        return jsonify({"error": str(e)}), 500


@re_bp.route("/compliance-activities/<int:comp_id>/control-activity", methods=["GET"])
def update_control_activity_by_comp_id(comp_id: int):
    """
    Updates or creates a ControlActivity for a given ComplianceActivity ID.
    It fetches the compliance activity and its associated clause, then uses an LLM
    to generate or update the ControlActivity and its related records.
    """
    try:
        # Note: All logic related to vec_id has been removed as per the user's request.

        with session_scope() as session:
            # Eagerly load the compliance activity and its related clause to avoid extra queries
            compliance_activity = (
                session.query(ComplianceActivities)
                .options(joinedload(ComplianceActivities.clauses))
                .filter(ComplianceActivities.id == comp_id)
                .first()
            )

            if not compliance_activity:
                return (
                    jsonify(
                        {"error": f"ComplianceActivity with ID {comp_id} not found"}
                    ),
                    404,
                )

            # Find or create the ControlActivity record
            control = (
                session.query(ControlActivity)
                .filter_by(compliance_activity_id=comp_id)
                .first()
            )

            if not control:
                # Create a new ControlActivity if it doesn't exist
                control = ControlActivity(
                    compliance_activity_id=comp_id,
                    activity_code=f"CA-{comp_id}",
                    activity_name=f"Control for Activity {comp_id}",
                )
                session.add(control)
                session.flush()
                flash("New Control Activity created", "info")

            # Prepare the data for the LLM call
            clause_text = compliance_activity.clauses.clause_text
            compliance_activity_payload = compliance_activity.to_dict()

            # Simulate LLM call to get updated data for the control activity
            updated_data = extract_structured_info_2(
                test_procedure(clause_text, compliance_activity_payload),
                ControlWorkpaper,
            )
            updated_data_dict = _as_dict(updated_data) or {}

            # Find or create the ControlActivity record
            control = (
                session.query(ControlActivity)
                .filter_by(compliance_activity_id=comp_id)
                .first()
            )

            if not control:
                # If no control activity exists, create a new one
                control = ControlActivity(compliance_activity_id=comp_id)
                session.add(control)
                session.flush()  # flush to get the ID for related records

            # --- Update the ControlActivity record itself ---
            control.activity_code = updated_data_dict.get("activity_code")
            control.activity_name = updated_data_dict.get("activity_name")
            control.activity_description = updated_data_dict.get("activity_description")
            control.objective = updated_data_dict.get("objective")
            control.owner = updated_data_dict.get("owner")
            control.control_type = updated_data_dict.get("control_type")
            control.frequency = updated_data_dict.get("frequency")
            control.sampling_guidance = updated_data_dict.get("sampling_guidance")
            control.auditor_observation = updated_data_dict.get("auditor_observation")
            control.findings = updated_data_dict.get("findings")
            control.impact = updated_data_dict.get("impact")
            control.severity = updated_data_dict.get("severity")
            control.recommendations = updated_data_dict.get("recommendations")
            control.reviewer_notes = updated_data_dict.get("reviewer_notes")
            control.explain_test_procedure = updated_data_dict.get(
                "explain_test_procedure"
            )

            # --- Handle related records (TestSteps, Interviews, etc.) ---
            # Find or create TestSteps
            test_steps_payload = updated_data_dict.get("test_procedure") or {}
            test_steps = control.test_procedure
            if not test_steps:
                test_steps = TestSteps(control_id=control.id)
                session.add(test_steps)
                session.flush()

            # Update TestSteps attributes in place
            test_steps.walkthrough = _ci_get(test_steps_payload, "walkthrough")
            test_steps.sampling = _ci_get(test_steps_payload, "sampling")

            # Sync Document Reviews
            new_docs_list = (
                test_steps_payload.get("review_of_documentation")
                or test_steps_payload.get("review_of_documents")
                or []
            )
            if isinstance(new_docs_list, (str, bytes)):
                new_docs_list = [new_docs_list]

            existing_doc_names = {doc.document_name for doc in test_steps.documents}
            new_doc_names = set(new_docs_list)

            # NOTE: We are no longer deleting documents to avoid foreign key violations.
            # This assumes that old, unused document review records can be left in the database.

            # Add new documents
            docs_to_add = new_doc_names - existing_doc_names
            for doc_name in docs_to_add:
                session.add(
                    DocumentReview(
                        test_procedure_id=test_steps.id, document_name=doc_name
                    )
                )

            # Sync Interviews
            interviews_data = test_steps_payload.get("interviews") or {}
            interview = test_steps.interviews
            if not interview:
                interview = Interview(test_procedure_id=test_steps.id)
                session.add(interview)
                session.flush()

            # Sync Interview Roles
            new_roles = set(interviews_data.get("roles") or [])
            existing_roles = {role.role for role in interview.roles}
            # NOTE: We are only adding new roles, not deleting old ones.
            roles_to_add = new_roles - existing_roles
            for role_name in roles_to_add:
                session.add(InterviewRole(interview_id=interview.id, role=role_name))

            # Sync Interview Questions
            new_questions = set(
                interviews_data.get("key_questions")
                or interviews_data.get("questions")
                or []
            )
            existing_questions = {q.question for q in interview.questions}
            # NOTE: We are only adding new questions, not deleting old ones.
            questions_to_add = new_questions - existing_questions
            for question_text in questions_to_add:
                session.add(
                    InterviewQuestion(interview_id=interview.id, question=question_text)
                )

            # Evidence artifacts: Clear existing and add new
            control.evidences.clear()
            evidence_input = updated_data_dict.get("evidences_artifacts_needed") or []
            if isinstance(evidence_input, dict):
                iter_evidence = (
                    (k, v if isinstance(v, list) else [v])
                    for k, v in evidence_input.items()
                )
            elif isinstance(evidence_input, list):

                def _yield_from_list(lst):
                    for entry in lst:
                        if not entry:
                            continue
                        if isinstance(entry, dict):
                            cat = (
                                entry.get("category") or entry.get("name") or "Unknown"
                            )
                            items = entry.get("items") or entry.get("items_list") or []
                            yield (cat, items if isinstance(items, list) else [items])
                        else:
                            yield ("Unknown", [str(entry)])

                iter_evidence = _yield_from_list(evidence_input)
            else:
                iter_evidence = []

            for category, items in iter_evidence:
                category = (category or "Unknown").strip()
                for item in items or []:
                    artifact = (
                        session.query(EvidenceArtifact)
                        .filter_by(category=category, item=item)
                        .first()
                    )
                    if not artifact:
                        artifact = EvidenceArtifact(category=category, item=item)
                        session.add(artifact)
                    control.evidences.append(artifact)

            session.flush()
            session.commit()
            flash("Control Activity Updated", "success")
            logger.info(
                "ControlActivity for ComplianceActivity ID %s updated/created successfully.",
                comp_id,
            )
            return redirect(request.referrer)

    except Exception as e:
        session.rollback()
        logger.error(
            "Error updating ControlActivity for Compliance ID %s: %s", comp_id, str(e)
        )
        print(e)
        flash(f"Error updating ControlActivity {e}", "error")
        return redirect(request.referrer)


@re_bp.route("/guidelines/<int:guideline_id>/generate_single_clause_activities")
def trigger_generate_missing_activities(guideline_id):
    """
    Auto-trigger cron job to generate activities for ONE clause without activities at a time
    URL: http://127.0.0.1:5000/re/guidelines/116/generate_single_clause_activities
    """
    try:
        # Check if guideline exists
        guideline = Guidelines.query.get(guideline_id)
        if not guideline:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": f"Guideline with ID {guideline_id} not found",
                    }
                ),
                404,
            )

        # Find the FIRST clause without activities (LIMIT 1)
        clause_without_activities = (
            Clauses.query.filter(
                Clauses.guideline_id == guideline_id,
                ~Clauses.id.in_(
                    db.session.query(ComplianceActivities.clause_id).filter(
                        ComplianceActivities.clause_id.isnot(None)
                    )
                ),
            )
            .order_by(Clauses.id.asc())
            .first()
        )  # Use first() instead of all()

        if not clause_without_activities:
            return f"""
            <html>
                <body style="font-family: Arial, sans-serif; text-align: center; padding: 50px;">
                    <h2 style="color: green;">✅ All clauses already have activities</h2>
                    <p>Guideline ID: {guideline_id}</p>
                    <p>All clauses in this guideline already have compliance activities.</p>
                    <a href="/re/clauses?guideline_id={guideline_id}">Back to Guidelines</a>
                </body>
            </html>
            """

        # Count remaining clauses for info display
        remaining_clauses_count = Clauses.query.filter(
            Clauses.guideline_id == guideline_id,
            ~Clauses.id.in_(
                db.session.query(ComplianceActivities.clause_id).filter(
                    ComplianceActivities.clause_id.isnot(None)
                )
            ),
        ).count()

        # FIX: Pass single clause ID (integer), not list
        task = generate_single_clause_activities.delay(
            guideline_id, clause_without_activities.id  # Pass integer, not list
        )

        return f"""
        <html>
            <head>
                <style>
                    .loading-spinner {{
                        border: 4px solid #f3f3f3;
                        border-top: 4px solid #3498db;
                        border-radius: 50%;
                        width: 40px;
                        height: 40px;
                        animation: spin 2s linear infinite;
                        margin: 20px auto;
                    }}
                    @keyframes spin {{
                        0% {{ transform: rotate(0deg); }}
                        100% {{ transform: rotate(360deg); }}
                    }}
                    .clause-info {{
                        background: #f8f9fa;
                        padding: 15px;
                        border-radius: 5px;
                        margin: 15px 0;
                    }}
                </style>
            </head>
            <body style="font-family: Arial, sans-serif; text-align: center; padding: 50px;">
                <div class="loading-spinner"></div>
                <h2 style="color: blue;">🔄 Activity Generation Started</h2>
                
                <div class="clause-info">
                    <p><strong>Processing Clause:</strong> {clause_without_activities.clause_no or 'N/A'}</p>
                    <p><strong>Clause ID:</strong> {clause_without_activities.id}</p>
                    <p><strong>Remaining Clauses:</strong> {remaining_clauses_count - 1}</p>
                </div>
                
                <p><strong>Guideline ID:</strong> {guideline_id}</p>
                <p><strong>Task ID:</strong> {task.id}</p>
                
                <p style="color: #666; margin-top: 20px;">
                    Generating activities for one clause. Hit this endpoint again to process the next clause.
                </p>
                
                <div style="margin-top: 20px;">
                    <a href="/re/clauses?guideline_id={guideline_id}" style="margin-right: 15px;">
                        Back to Guidelines
                    </a>
                    <a href="/re/guidelines/{guideline_id}/generate_single_clause_activities" 
                       style="background: #007bff; color: white; padding: 10px 15px; text-decoration: none; border-radius: 4px;">
                        Process Next Clause
                    </a>
                </div>
            </body>
        </html>
        """

    except Exception as e:
        return (
            f"""
        <html>
            <body style="font-family: Arial, sans-serif; text-align: center; padding: 50px;">
                <h2 style="color: red;">❌ Error</h2>
                <p>Failed to trigger activity generation: {str(e)}</p>
                <p><strong>Guideline ID:</strong> {guideline_id}</p>
                <a href="/re/clauses?guideline_id={guideline_id}">Back to Guidelines</a>
            </body>
        </html>
        """,
            500,
        )


@re_bp.route("/retry-pending-activities", methods=["POST"])
@role_required("COMPLIFYRE", "AUDITOR", "RE")
def retry_pending_activities():
    """Retry activity generation for clauses without activities."""
    try:
        data = request.get_json()
        guideline_id = data.get("guideline_id")
        if not guideline_id:
            return jsonify({"status": "error", "message": "guideline_id required"}), 400

        # Find clauses without activities
        clauses = Clauses.query.filter_by(guideline_id=guideline_id).all()
        pending_clauses = [
            c for c in clauses
            if not c.compliance_activities or len(c.compliance_activities) == 0
        ]

        if not pending_clauses:
            return jsonify({
                "status": "success",
                "message": "No pending clauses found",
                "triggered_count": 0
            })

        # Trigger activities generation for each pending clause
        from app.services.manual_task import extract_activities
        triggered_count = 0
        for clause in pending_clauses:
            extract_activities.apply_async(
                args=[clause.id],
                queue='extract_activities'
            )
            triggered_count += 1

        logger.info(
            f"[RETRY] Triggered activities for {triggered_count} pending clauses "
            f"in guideline_id={guideline_id}"
        )

        return jsonify({
            "status": "success",
            "message": f"Triggered activity generation for {triggered_count} pending clauses",
            "triggered_count": triggered_count,
            "guideline_id": guideline_id,
        })

    except Exception as e:
        logger.exception(f"Error retrying pending activities: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
